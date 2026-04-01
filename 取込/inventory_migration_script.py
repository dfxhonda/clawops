#!/usr/bin/env python3
"""
棚卸し移行スクリプト: 景品棚卸リスト 3期 → システム移行表
"""
import openpyxl, json, re, unicodedata
from datetime import datetime
from openpyxl.styles import Font, PatternFill

INPUT_FILE = '取込/景品棚卸リスト　3期.xlsx'
MASTERS_FILE = '/tmp/prize_masters.json'
OUTPUT_DRAFT = '取込/inventory_migration_draft.xlsx'
OUTPUT_REVIEW = '取込/inventory_review_needed.xlsx'
OUTPUT_DICT = '取込/inventory_name_mapping_dictionary.xlsx'
OUTPUT_LOG = '取込/inventory_migration_log.txt'

GACHA_KEYWORDS = [
    'スイッチ', 'Switch', 'PS5', 'ps5', 'ポータル', 'プレステ',
    'NIKE', 'ナイキ', 'スニーカー', 'リファ', 'refa',
    'ブランドコスメ', 'シュプリーム', 'DW', 'UGG',
    'ポケカ', 'fire', 'Fire', 'instax', 'iPad', 'ipad',
    'ラブブBOX', 'ラブブエナジー',
    'A景品', 'カプセル景品', '機械内景品',
    'PSポータル', 'リモートプレーヤー', '神袋',
]
GACHA_PRICE_THRESHOLD = 5000

log_lines = []
def log(msg):
    log_lines.append(msg)
    print(msg)

def normalize(s):
    if not s: return ''
    s = str(s).strip()
    s = unicodedata.normalize('NFKC', s)
    s = s.lower()
    s = re.sub(r'[\s\u3000]+', '', s)
    s = re.sub(r'[（）\(\)\[\]【】「」『』\-_・/]', '', s)
    return s

def is_gacha_item(name, price):
    if not name: return False
    for kw in GACHA_KEYWORDS:
        if kw.lower() in name.lower() or kw in name:
            return True
    try:
        if float(price or 0) >= GACHA_PRICE_THRESHOLD:
            return True
    except: pass
    return False

# マスタ読み込み
log(f"[{datetime.now().isoformat()}] 移行処理開始")
with open(MASTERS_FILE) as f:
    masters = json.load(f)
log(f"prize_masters: {len(masters)}件")

master_index = []
for m in masters:
    names = [m['prize_name']]
    aliases = m.get('aliases', '[]') or '[]'
    try:
        al = json.loads(aliases) if isinstance(aliases, str) else aliases
        if isinstance(al, list): names.extend(al)
    except: pass
    for n in names:
        if n:
            master_index.append((normalize(n), m['prize_id'], m['prize_name'], m.get('original_cost'), m.get('status')))
log(f"検索インデックス: {len(master_index)}エントリ")

def find_candidates(short_name, price=None):
    if not short_name or not short_name.strip(): return []
    norm = normalize(short_name)
    if len(norm) < 2: return []
    scores = []
    for idx_norm, pid, pname, pcost, pstatus in master_index:
        score = 0
        reason = []
        if norm == idx_norm:
            score = 100; reason.append('完全一致')
        elif norm in idx_norm:
            score = 70 + min(len(norm), 20); reason.append('短縮名⊂マスタ名')
        elif idx_norm in norm:
            score = 60 + min(len(idx_norm), 15); reason.append('マスタ名⊂短縮名')
        else:
            best = 0
            for s in range(len(norm)):
                for e in range(s+3, min(s+30, len(norm)+1)):
                    if norm[s:e] in idx_norm: best = max(best, e-s)
            if best >= 3:
                score = min(55, best * 4); reason.append(f'部分一致({best}文字)')
        if score > 0 and price and pcost:
            try:
                p1, p2 = float(price), float(pcost)
                if p1 > 0 and p2 > 0 and min(p1,p2)/max(p1,p2) > 0.8:
                    score += 10; reason.append('単価近似')
            except: pass
        if score > 0:
            scores.append({'prize_id':pid,'prize_name':pname,'score':score,'reason':'・'.join(reason),'original_cost':pcost,'status':pstatus})
    scores.sort(key=lambda x:-x['score'])
    seen = set(); result = []
    for s in scores:
        if s['prize_id'] not in seen:
            seen.add(s['prize_id']); result.append(s)
            if len(result) >= 3: break
    return result

def score_to_level(score):
    if score >= 80: return 'High'
    if score >= 50: return 'Medium'
    return 'Low'

# Excel解析
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
all_rows = []

for sname in wb.sheetnames:
    if sname == '合計': continue
    ws = wb[sname]
    log(f"\nシート '{sname}' 解析中... (行数:{ws.max_row}, 列数:{ws.max_column})")

    # 右半分検出（2期分の場合）
    col_offset = 0
    if sname in ('中島', '久留米 '):
        row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
        for ci in range(5, len(row2)):
            v = str(row2[ci] or '')
            if '店舗' in v or '景品' in v:
                col_offset = ci; break
        if col_offset:
            log(f"  2期分検出。右側(列{col_offset+1}〜)を使用")

    # ヘッダー解析
    h2 = [str(ws.cell(row=2, column=c).value or '').strip() for c in range(1, ws.max_column+1)]
    h3 = [str(ws.cell(row=3, column=c).value or '').strip() for c in range(1, ws.max_column+1)]

    store_col = name_col = machine_col = price_col = None
    qty_cols = []

    for i, h in enumerate(h2):
        if i < col_offset: continue
        if '店舗' in h and store_col is None: store_col = i
        elif '景品' in h and name_col is None: name_col = i
        elif '機械' in h and machine_col is None: machine_col = i
        elif '単価' in h and price_col is None: price_col = i

    for i, h in enumerate(h3):
        if i < col_offset: continue
        if '倉庫' in h or '手持ち' in h or '機械' in h:
            qty_cols.append((i, h))

    if name_col is None:
        log(f"  ⚠ 景品名列が見つからない。スキップ")
        continue

    log(f"  列: store={store_col} name={name_col} machine={machine_col} price={price_col} qty={qty_cols}")

    cur_store = sname
    cur_machine = ''
    count = 0

    for ri in range(5 if sname != '久留米(ドンキ景品)' else 4, ws.max_row + 1):
        cells = [ws.cell(row=ri, column=c+1).value for c in range(ws.max_column)]

        if store_col is not None and cells[store_col]:
            sv = str(cells[store_col]).strip()
            if sv and sv != '\u3000': cur_store = sv

        if machine_col is not None and cells[machine_col]:
            mv = str(cells[machine_col]).strip()
            if mv and mv != '\u3000': cur_machine = mv

        pn = cells[name_col]
        if not pn or str(pn).strip() in ('', '\u3000'): continue
        pn = str(pn).strip()

        pr = cells[price_col] if price_col is not None else None

        total_qty = 0
        for qi, ql in qty_cols:
            v = cells[qi]
            try: n = int(float(v)) if v and str(v).strip() not in ('', '\u3000') else 0
            except: n = 0
            total_qty += n

        all_rows.append({
            'sheet': sname, 'store': cur_store, 'machine': cur_machine,
            'prize_name': pn, 'price': pr, 'total_qty': total_qty,
        })
        count += 1

    log(f"  → {count}行抽出")

log(f"\n全シート合計: {len(all_rows)}行")

# マッチング
log("\nマッチング処理開始...")
results = []
for row in all_rows:
    name = row['prize_name']
    price = row['price']
    gacha = is_gacha_item(name, price)
    candidates = find_candidates(name, price) if not gacha else []

    if gacha:
        status = '要確認'; level = ''; note = '高額ガチャ/ロッカー/手買い候補'
    elif not candidates:
        status = '不明'; level = ''; note = '候補なし'
    elif len(candidates) == 1 and candidates[0]['score'] >= 80:
        status = '一致'; level = 'High'; note = candidates[0]['reason']
    elif candidates[0]['score'] >= 60:
        if len(candidates) > 1 and candidates[1]['score'] > 40:
            status = '要確認'; level = score_to_level(candidates[0]['score']); note = candidates[0]['reason'] + ' (複数候補)'
        else:
            status = '一致'; level = score_to_level(candidates[0]['score']); note = candidates[0]['reason']
    else:
        status = '要確認'; level = score_to_level(candidates[0]['score']) if candidates else ''
        note = (candidates[0]['reason'] if candidates else '') + ' (低一致度)'

    loc_type = '拠点倉庫' if row['sheet'] in ('久留米 ', '久留米(ドンキ景品)', '飯塚', '鹿児島') else '担当者車'
    results.append({**row, 'location_type': loc_type, 'candidates': candidates, 'gacha': gacha, 'status': status, 'level': level, 'note': note})

total = len(results)
matched = sum(1 for r in results if r['status'] == '一致')
review = sum(1 for r in results if r['status'] == '要確認')
unknown = sum(1 for r in results if r['status'] == '不明')
gacha_count = sum(1 for r in results if r['gacha'])
log(f"\n─── 集計 ───")
log(f"総行数: {total}")
log(f"一致: {matched}")
log(f"要確認: {review}")
log(f"不明: {unknown}")
log(f"高額ガチャ/ロッカー/手買い候補: {gacha_count}")

# 出力
hfill = PatternFill(start_color='2E2E3E', end_color='2E2E3E', fill_type='solid')
hfont = Font(bold=True, color='E8E8F0')

# 1. Draft
log(f"\n{OUTPUT_DRAFT} 作成中...")
wb1 = openpyxl.Workbook(); ws1 = wb1.active; ws1.title = '移行表'
ws1.append(['拠点','店舗','場所区分','元の短縮景品名','機械名','単価','数量',
    '候補景品名1','候補景品名2','候補景品名3','候補景品ID1','候補景品ID2','候補景品ID3',
    '一致度','確認状態','最終採用景品ID','最終採用景品名','備考'])
for r in results:
    c = r['candidates']
    fid = c[0]['prize_id'] if r['status']=='一致' and r['level']=='High' and c else ''
    fn = c[0]['prize_name'] if r['status']=='一致' and r['level']=='High' and c else ''
    ws1.append([r['sheet'],r['store'],r['location_type'],r['prize_name'],r['machine'],r['price'],r['total_qty'],
        c[0]['prize_name'] if len(c)>0 else '',c[1]['prize_name'] if len(c)>1 else '',c[2]['prize_name'] if len(c)>2 else '',
        c[0]['prize_id'] if len(c)>0 else '',c[1]['prize_id'] if len(c)>1 else '',c[2]['prize_id'] if len(c)>2 else '',
        r['level'],r['status'],fid,fn,r['note']])
for cell in ws1[1]: cell.fill=hfill; cell.font=hfont
for col,w in [('A',12),('D',30),('H',30),('I',30),('J',30),('O',10),('Q',30)]:
    ws1.column_dimensions[col].width = w
wb1.save(OUTPUT_DRAFT)
log(f"  → {total}行")

# 2. Review
log(f"\n{OUTPUT_REVIEW} 作成中...")
wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = '要確認'
ws2.append(['拠点','店舗','場所区分','元の短縮景品名','機械名','単価','数量',
    '候補景品名1','候補景品名2','候補景品名3','要確認理由','備考'])
rev = [r for r in results if r['status'] in ('要確認','不明')]
for r in rev:
    c = r['candidates']
    reason = []
    if r['gacha']: reason.append('高額ガチャ/手買い')
    if not c: reason.append('候補なし')
    elif len(c)>1 and c[1]['score']>40: reason.append('複数候補')
    if r['level']=='Low': reason.append('低一致度')
    if not reason: reason.append(r['note'])
    ws2.append([r['sheet'],r['store'],r['location_type'],r['prize_name'],r['machine'],r['price'],r['total_qty'],
        c[0]['prize_name'] if len(c)>0 else '',c[1]['prize_name'] if len(c)>1 else '',c[2]['prize_name'] if len(c)>2 else '',
        '／'.join(reason),r['note']])
for cell in ws2[1]: cell.fill=hfill; cell.font=hfont
wb2.save(OUTPUT_REVIEW)
log(f"  → {len(rev)}行")

# 3. Dictionary
log(f"\n{OUTPUT_DICT} 作成中...")
wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = '対応辞書'
ws3.append(['拠点','場所区分','元の短縮景品名','採用景品ID','採用景品名','確認状態','備考'])
seen = set(); dc = 0
for r in results:
    if r['status']=='一致' and r['candidates'] and r['prize_name'] not in seen:
        seen.add(r['prize_name']); c=r['candidates'][0]
        ws3.append([r['sheet'],r['location_type'],r['prize_name'],c['prize_id'],c['prize_name'],r['status'],r['note']])
        dc += 1
for cell in ws3[1]: cell.fill=hfill; cell.font=hfont
wb3.save(OUTPUT_DICT)
log(f"  → {dc}件")

# 紛らわしい例
log("\n─── 紛らわしかった短縮名 ───")
amb = [r for r in results if r['status']=='要確認' and len(r['candidates'])>=2]
for r in amb[:10]:
    c=r['candidates']
    log(f"  「{r['prize_name']}」→ {c[0]['prize_name'][:25]}({c[0]['score']}) / {c[1]['prize_name'][:25]}({c[1]['score']})")

with open(OUTPUT_LOG, 'w') as f: f.write('\n'.join(log_lines))
log(f"\n─── 出力完了 ───")
for f in [OUTPUT_DRAFT, OUTPUT_REVIEW, OUTPUT_DICT, OUTPUT_LOG]:
    log(f"  {f}")
