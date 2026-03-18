#!/usr/bin/env python3
"""
景品フォームCSV + (株)Change.xlsx → ClawOps景品マスタCSV変換スクリプト

出力:
  prizes_master.csv  - 景品マスタ（商品コードで重複排除）
  orders_all.csv     - 全発注履歴（重複あり）
"""
import csv
import glob
import os
import re
import math
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    exit(1)

FOLDER = os.path.dirname(os.path.abspath(__file__))

# ============================================
# 商品名からの情報抽出
# ============================================
def extract_from_name(name):
    """商品名から短縮名・サイズ・入数・@単価を抽出"""
    if not name:
        return {'short_name': '', 'item_size': '', 'pcs_from_name': 0, 'at_price': 0}

    # --- 入数抽出 (XX入) ---
    pcs_from_name = 0
    m = re.search(r'(\d+)入', name)
    if m:
        pcs_from_name = int(m.group(1))

    # --- @単価抽出 ---
    at_price = 0
    m2 = re.search(r'[@＠](\d+\.?\d*)', name)
    if m2:
        at_price = float(m2.group(1))

    # --- サイズ抽出 ---
    item_size = ''
    size_patterns = [
        r'\b(BIG|big|Big)\b',
        r'\b(XXL|XL|LL|L|M|S|SS)\b',
        r'(\d+[cm×xX]\d+[cm]*)',
        r'(\d+cm)',
        r'(特大|大|中|小|ミニ)',
        r'(Lサイズ|Mサイズ|Sサイズ)',
        r'(ぬいぐるみ|メガ|ギガ|ジャンボ)',
    ]
    for pat in size_patterns:
        m3 = re.search(pat, name, re.IGNORECASE)
        if m3:
            item_size = m3.group(1) or m3.group(0)
            break

    # --- サイズをcm表記からも抽出 ---
    if not item_size:
        m4 = re.search(r'(\d+)㎝', name)
        if m4:
            item_size = m4.group(1) + 'cm'

    # --- 短縮名 ---
    short_name = name
    # 括弧・サイズ表記・付帯情報を除去
    short_name = re.sub(r'【.*?】', '', short_name)
    short_name = re.sub(r'\(.*?\)', '', short_name)
    short_name = re.sub(r'（.*?）', '', short_name)
    short_name = re.sub(r'\d+入.*$', '', short_name)  # XX入以降を削除
    short_name = re.sub(r'[@＠]\d+.*$', '', short_name)  # @XX以降を削除
    short_name = re.sub(r'\s*(BIG|XXL|XL|LL|Lサイズ|Mサイズ|Sサイズ|特大|ジャンボ|メガ|ギガ)\s*', ' ', short_name, flags=re.IGNORECASE)
    short_name = re.sub(r'\s+', ' ', short_name).strip()
    if len(short_name) > 20:
        short_name = short_name[:20]

    return {
        'short_name': short_name,
        'item_size': item_size,
        'pcs_from_name': pcs_from_name,
        'at_price': at_price,
    }


def map_category(raw_cat):
    """景品フォームのカテゴリー → ClawOpsカテゴリにマッピング"""
    if not raw_cat:
        return ''
    raw = raw_cat.strip()
    # 複数カテゴリが入ることがある（カンマ区切り等）→ 最初の有効なものを採用
    mapping = {
        'ぬいぐるみ': 'ぬいぐるみ',
        'お菓子': '食品',
        '食品': '食品',
        '電子商品': '雑貨',
        '備品': '雑貨',
        'SNS・流行り物': '雑貨',
        'ガチャ景品': '雑貨',
        '箱物系': '雑貨',
        '大型機用': '',
        '小型機用': '',
        '即納': '',
        '特価品': '',
        'オススメ商品': '',
        'クォーター': '',
        '送料無料': '',
    }
    for key, val in mapping.items():
        if key in raw:
            if val:
                return val
    return ''


def parse_date_from_noki(noki_str):
    """納期文字列 (例: "2026年03月 上旬予定") → YYYY-MM-DD (月末)"""
    if not noki_str:
        return ''
    m = re.search(r'(\d{4})年(\d{1,2})月', noki_str)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        # 月末日を返す
        if mo == 12:
            return f'{y}-12-31'
        from calendar import monthrange
        _, last_day = monthrange(y, mo)
        return f'{y}-{mo:02d}-{last_day:02d}'
    return ''


# ============================================
# 景品フォームCSV読み込み
# ============================================
def load_form_csvs():
    """景品フォームCSV全ファイルを読み込み"""
    csvfiles = sorted(glob.glob(os.path.join(FOLDER, '*景品フォーム発注履歴CSV*.csv')))
    print(f"景品フォームCSV: {len(csvfiles)}ファイル")

    all_orders = []
    products = {}  # 商品コード → 最新データ

    for fpath in csvfiles:
        # ファイル名から伝票月を抽出
        fname = os.path.basename(fpath)
        slip_month = ''
        m = re.search(r'伝票月：(\d{4}年\d{2}月)', fname)
        if m:
            slip_month = m.group(1)

        with open(fpath, encoding='utf-8-sig') as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                name = (row.get('商品名') or '').strip()
                if not name:
                    continue

                code = (row.get('商品コード') or '').strip()
                cost_str = (row.get('単価(税別)') or '').replace(',', '')
                pcs_str = (row.get('入数') or '').replace(',', '')
                qty_str = (row.get('受注数') or '').replace(',', '')
                total_str = (row.get('合計金額(税別)') or '').replace(',', '')
                cat = (row.get('カテゴリー') or '').strip()
                noki = (row.get('納期') or '').strip()
                order_dt = (row.get('発注日時') or '').strip()
                ship_status = (row.get('発送状況') or '').strip()
                dest = (row.get('発送先') or '').strip()

                # 数値変換
                try:
                    cost = float(cost_str) if cost_str else 0
                except:
                    cost = 0
                try:
                    pcs = float(pcs_str) if pcs_str else 0
                except:
                    pcs = 0
                try:
                    qty = float(qty_str) if qty_str else 0
                except:
                    qty = 0
                try:
                    total = float(total_str) if total_str else 0
                except:
                    total = 0

                # 名前から入数を補完
                extracted = extract_from_name(name)
                real_pcs = pcs
                if pcs <= 1 and extracted['pcs_from_name'] > 1:
                    real_pcs = extracted['pcs_from_name']

                # 1個単価計算
                if extracted['at_price'] > 0:
                    unit_price = round(extracted['at_price'])
                elif real_pcs > 0 and qty > 0 and total > 0:
                    unit_price = round(total / (qty * real_pcs))
                elif cost > 0 and real_pcs > 1:
                    unit_price = round(cost / real_pcs)
                else:
                    unit_price = round(cost)

                # 発注日
                order_date = ''
                if order_dt:
                    m = re.match(r'(\d{4}-\d{2}-\d{2})', order_dt)
                    if m:
                        order_date = m.group(1)

                # 到着予定日
                arrival_date = parse_date_from_noki(noki)

                order_record = {
                    'code': code,
                    'prize_name': name,
                    'short_name': extracted['short_name'],
                    'item_size': extracted['item_size'],
                    'unit_cost': unit_price,
                    'category': map_category(cat),
                    'raw_category': cat,
                    'supplier_name': '景品フォーム',
                    'jan_code': code,
                    'order_at': order_date,
                    'arrival_at': arrival_date,
                    'case_count': int(qty) if qty == int(qty) else qty,
                    'pieces_per_case': int(real_pcs) if real_pcs == int(real_pcs) else real_pcs,
                    'total_cost': round(total),
                    'slip_month': slip_month,
                    'dest': dest,
                }

                all_orders.append(order_record)

                # 商品マスタ: 商品コードで最新を保持
                if code:
                    if code not in products or order_date > products[code].get('order_at', ''):
                        products[code] = order_record.copy()

    return products, all_orders


# ============================================
# (株)Change xlsx読み込み
# ============================================
def load_change_xlsx():
    """(株)Change.xlsx の全シートを読み込み"""
    products = {}
    all_orders = []

    xlsx_files = glob.glob(os.path.join(FOLDER, '*Change*.xlsx'))
    print(f"Change xlsx: {len(xlsx_files)}ファイル")

    for fpath in xlsx_files:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))

            # ヘッダー行を探す（"商品名"が含まれる行）
            header_idx = None
            for i, row in enumerate(rows):
                if row and any(str(c or '').strip() == '商品名' for c in row):
                    header_idx = i
                    break

            if header_idx is None:
                continue

            # データ行を読む
            for row in rows[header_idx + 1:]:
                if not row or not row[0]:
                    continue
                name = str(row[0] or '').strip()
                if not name or name.startswith('下記') or name.startswith('合計'):
                    continue

                cases = 0
                pcs = 0
                unit_cost = 0
                note = ''
                dest = ''

                try:
                    cases = int(row[1]) if row[1] else 0
                except:
                    pass
                try:
                    pcs = int(row[3]) if row[3] else 0
                except:
                    pass
                try:
                    unit_cost = int(row[4]) if row[4] else 0
                except:
                    pass
                try:
                    note = str(row[7] or '').strip()  # 備考=納期
                except:
                    pass
                try:
                    dest = str(row[8] or '').strip()
                except:
                    pass

                extracted = extract_from_name(name)
                code = f'CHG-{sname}-{name[:10]}'

                order_record = {
                    'code': code,
                    'prize_name': name,
                    'short_name': extracted['short_name'],
                    'item_size': extracted['item_size'],
                    'unit_cost': unit_cost,
                    'category': '',
                    'raw_category': '',
                    'supplier_name': '(株)Change',
                    'jan_code': '',
                    'order_at': '',
                    'arrival_at': '',
                    'case_count': cases,
                    'pieces_per_case': pcs,
                    'total_cost': cases * pcs * unit_cost if cases and pcs else 0,
                    'slip_month': sname,
                    'dest': dest,
                }

                all_orders.append(order_record)

                if name not in products:
                    products[name] = order_record.copy()

        wb.close()

    return products, all_orders


# ============================================
# メイン処理
# ============================================
def main():
    print("=" * 60)
    print("ClawOps 景品データ変換")
    print("=" * 60)

    # 読み込み
    form_products, form_orders = load_form_csvs()
    change_products, change_orders = load_change_xlsx()

    print(f"\n景品フォーム: {len(form_products)}商品 / {len(form_orders)}発注行")
    print(f"Change: {len(change_products)}商品 / {len(change_orders)}発注行")

    # 統合
    all_products = {}
    for code, p in form_products.items():
        all_products[code] = p
    for name, p in change_products.items():
        key = f'CHG_{name}'
        if key not in all_products:
            all_products[key] = p

    all_orders = form_orders + change_orders

    print(f"統合: {len(all_products)}商品 / {len(all_orders)}発注行")

    # --- prizes_master.csv 出力 ---
    master_path = os.path.join(FOLDER, 'prizes_master.csv')
    master_headers = [
        '商品名', '短縮名', 'サイズ', 'カテゴリ', '単価',
        'サプライヤー', 'JANコード', '発注日', '到着予定日',
        'ケース数', '入数', '在庫数', 'ステータス',
    ]
    with open(master_path, 'w', encoding='utf-8-sig', newline='') as fp:
        writer = csv.writer(fp)
        writer.writerow(master_headers)
        for key, p in sorted(all_products.items(), key=lambda x: x[1].get('order_at', ''), reverse=True):
            writer.writerow([
                p['prize_name'],
                p['short_name'],
                p['item_size'],
                p['category'],
                p['unit_cost'],
                p['supplier_name'],
                p['jan_code'],
                p['order_at'],
                p['arrival_at'],
                p['case_count'],
                p['pieces_per_case'],
                0,       # 在庫数（不明）
                'FALSE',  # デフォルト非アクティブ
            ])

    print(f"\n✅ prizes_master.csv: {len(all_products)}行 → {master_path}")

    # --- orders_all.csv 出力 ---
    orders_path = os.path.join(FOLDER, 'orders_all.csv')
    order_headers = [
        '商品コード', '商品名', '短縮名', '単価', 'カテゴリ',
        'サプライヤー', '発注日', '到着予定日', 'ケース数', '入数',
        '合計金額', '伝票月', '配送先',
    ]
    with open(orders_path, 'w', encoding='utf-8-sig', newline='') as fp:
        writer = csv.writer(fp)
        writer.writerow(order_headers)
        for o in sorted(all_orders, key=lambda x: x.get('order_at', ''), reverse=True):
            writer.writerow([
                o['code'],
                o['prize_name'],
                o['short_name'],
                o['unit_cost'],
                o['raw_category'] or o['category'],
                o['supplier_name'],
                o['order_at'],
                o['arrival_at'],
                o['case_count'],
                o['pieces_per_case'],
                o['total_cost'],
                o['slip_month'],
                o['dest'],
            ])

    print(f"✅ orders_all.csv: {len(all_orders)}行 → {orders_path}")

    # --- サマリー ---
    print(f"\n--- サマリー ---")
    cats = {}
    for p in all_products.values():
        c = p['category'] or '未分類'
        cats[c] = cats.get(c, 0) + 1
    for c, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}商品")

    suppliers = {}
    for p in all_products.values():
        s = p['supplier_name']
        suppliers[s] = suppliers.get(s, 0) + 1
    print(f"\nサプライヤー別:")
    for s, n in sorted(suppliers.items(), key=lambda x: -x[1]):
        print(f"  {s}: {n}商品")

    # 単価 0円のチェック
    zero_price = sum(1 for p in all_products.values() if p['unit_cost'] == 0)
    print(f"\n単価0円: {zero_price}商品")


if __name__ == '__main__':
    main()
