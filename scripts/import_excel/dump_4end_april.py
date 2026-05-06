#!/usr/bin/env python3
"""4末付データ取込: Excel構造ダンプ (作業2)

Output: ~/clawops/imports/dump_2026_04_end.json + dump_2026_04_end.md
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

CANDIDATES = [
    Path.home() / "clawops/docs/lists/景品棚卸リスト　3期 (1).xlsx",
    Path.home() / "clawops/archive/取込/景品棚卸リスト　3期.xlsx",
]

OUT_DIR = Path.home() / "clawops/imports"
OUT_DIR.mkdir(parents=True, exist_ok=True)
JSON_OUT = OUT_DIR / "dump_2026_04_end.json"
MD_OUT = OUT_DIR / "dump_2026_04_end.md"


def cell_repr(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def dump_sheet(wb_path: Path, sheet_name: str) -> dict:
    wb = openpyxl.load_workbook(str(wb_path), data_only=True, read_only=False)
    ws = wb[sheet_name]
    merged = [str(r) for r in ws.merged_cells.ranges]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    head_rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(8, max_row), values_only=True):
        head_rows.append([cell_repr(c) for c in row])

    try:
        df = pd.read_excel(wb_path, sheet_name=sheet_name, header=None, nrows=200)
        dtypes = {f"col_{i}": str(df[i].dtype) for i in df.columns}
        non_null_counts = {f"col_{i}": int(df[i].notna().sum()) for i in df.columns}
    except Exception as e:
        dtypes = {"_error": str(e)}
        non_null_counts = {}

    return {
        "sheet": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "merged_cells": merged[:30],
        "merged_cells_total": len(merged),
        "head_rows_first8": head_rows,
        "dtypes_first200rows": dtypes,
        "non_null_counts_first200rows": non_null_counts,
    }


def dump_workbook(path: Path) -> dict:
    if not path.exists():
        return {"path": str(path), "error": "FILE_NOT_FOUND"}
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    out = {
        "path": str(path),
        "size_bytes": path.stat().st_size,
        "mtime_iso": __import__("datetime").datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        "sheet_names": sheets,
        "sheets": [],
    }
    wb.close()
    for s in sheets:
        out["sheets"].append(dump_sheet(path, s))
    return out


def md_section(book: dict) -> str:
    lines = []
    lines.append(f"### {book['path']}")
    if book.get("error"):
        lines.append(f"- ⚠️ {book['error']}")
        return "\n".join(lines)
    lines.append(f"- size: {book['size_bytes']:,} bytes / mtime: {book['mtime_iso']}")
    lines.append(f"- sheet_names ({len(book['sheet_names'])}): {book['sheet_names']}")
    for s in book["sheets"]:
        lines.append("")
        lines.append(f"#### sheet: {s['sheet']}")
        lines.append(f"- max_row x max_col: {s['max_row']} x {s['max_col']}")
        lines.append(f"- merged_cells_total: {s['merged_cells_total']} (上位30: {s['merged_cells'][:8]}...)")
        lines.append("- head rows (1-8):")
        for i, r in enumerate(s["head_rows_first8"], start=1):
            lines.append(f"  - row{i}: {r}")
        lines.append("- dtypes (first200rows):")
        lines.append(f"  - {s['dtypes_first200rows']}")
        lines.append(f"- non_null_counts (first200rows): {s['non_null_counts_first200rows']}")
    return "\n".join(lines)


def main() -> int:
    results = [dump_workbook(p) for p in CANDIDATES]
    JSON_OUT.write_text(json.dumps(results, ensure_ascii=False, indent=2))
    md = ["# 4末付データ取込: Excel構造ダンプ (作業2)", ""]
    md.append("候補ファイル:")
    for p in CANDIDATES:
        md.append(f"- {p}")
    md.append("")
    for book in results:
        md.append(md_section(book))
        md.append("\n---\n")
    MD_OUT.write_text("\n".join(md))
    print(f"OK wrote {JSON_OUT}")
    print(f"OK wrote {MD_OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
