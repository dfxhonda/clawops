# scripts/import_excel/phase1_masters.py
"""
Generates SQL for:
  1. New stores (天文館 + rental clients)
  2. New machines (those not already in DB by normalized name match)
  3. New booths (for new machines; existing machines' booths already present)

Usage: python phase1_masters.py
Outputs SQL to stdout. Execute via Supabase MCP.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from config import (DONKI_FILE, FUKUSHIGE_FILE, STORE_MAP, DONKI_SHEETS,
                    NEW_STORES, DEFAULT_OPERATOR_ID, IMPORT_TAG,
                    EXISTING_MACHINE_MAX)
from helpers import normalize, next_machine_code, booth_code, sq, new_id

# --- known existing machines: (store_code, normalized_name) -> machine_code ---
EXISTING_MACHINES = {
    ('KOS01', 'BUZZ#1'):    'KOS01-M02',
    ('KOS01', 'BUZZ#2'):    'KOS01-M03',
    ('KOS01', 'セサミ#3'):  'KOS01-M04',
    ('KOS01', 'セサミ#4'):  'KOS01-M05',
    ('KOS01', 'セサミ#1'):  'KOS01-M06',
    ('KOS01', 'セサミ#2'):  'KOS01-M07',
    ('KKY01', 'BUZZスリム'): 'KKY01-M01',
    ('KKY01', 'BUZZ#1'):    'KKY01-M02',
    ('MNK01', 'BUZZ#1'):    'MNK01-M01',
    ('MNK01', 'BUZZ#2'):    'MNK01-M02',
}

# existing booth counts per machine_code
EXISTING_BOOTHS = {
    'KOS01-M02': 4, 'KOS01-M03': 4,
    'KOS01-M04': 2, 'KOS01-M05': 2, 'KOS01-M06': 2, 'KOS01-M07': 2,
    'KKY01-M01': 2, 'KKY01-M02': 4,
    'MNK01-M01': 4, 'MNK01-M02': 4,
}


def parse_donki_machines():
    """Returns list of (store_code, machine_name_raw, booth_num)."""
    wb = openpyxl.load_workbook(DONKI_FILE, data_only=True)
    results = []
    for sheet_name in DONKI_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 1  # skip header row
        while i < len(rows):
            row = rows[i]
            if len(row) >= 2 and row[1] == '100円' and row[0] and str(row[0]).strip():
                store_name = str(row[0]).strip()
                machine_name = None
                booth_num = 1
                if i + 1 < len(rows) and rows[i+1][0]:
                    machine_name = str(rows[i+1][0]).strip()
                if i + 2 < len(rows) and rows[i+2][0] is not None:
                    try:
                        booth_num = int(rows[i+2][0])
                    except (ValueError, TypeError):
                        booth_num = 1
                store_code = STORE_MAP.get(store_name)
                if store_code and machine_name:
                    results.append((store_code, machine_name, booth_num))
                i += 12
            else:
                i += 1
    return results


def parse_fukushige_machines():
    """Returns list of (store_code, machine_name_raw, 1) — 1 booth per machine."""
    wb = openpyxl.load_workbook(FUKUSHIGE_FILE, data_only=True)
    ws = wb['福重']
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    results = []
    seen = set()
    for v in row1:
        if v and '福重' in str(v):
            full_name = str(v).strip()
            if full_name in seen:
                continue
            seen.add(full_name)
            machine_name = full_name.replace('福重', '', 1).strip()
            if machine_name:
                results.append(('FKS01', machine_name, 1))
    return results


def build_machine_registry(tuples):
    """
    Returns:
      new_machines: list of (machine_code, store_code, machine_name_raw)
      new_booths:   list of (booth_code, machine_code, store_code, booth_num)
    """
    m_state = dict(EXISTING_MACHINE_MAX)
    decided = {}  # (store_code, norm_name) -> machine_code

    new_machines = []
    new_booths = []

    for store_code, machine_name_raw, booth_num in tuples:
        norm = normalize(machine_name_raw)
        key = (store_code, norm)

        # Already in DB?
        if key in EXISTING_MACHINES:
            mc = EXISTING_MACHINES[key]
            existing_count = EXISTING_BOOTHS.get(mc, 0)
            if booth_num > existing_count:
                bc = booth_code(mc, booth_num)
                new_booths.append((bc, mc, store_code, booth_num))
            continue

        # Already decided in this run?
        if key in decided:
            mc = decided[key]
        else:
            mc = next_machine_code(store_code, m_state)
            decided[key] = mc
            new_machines.append((mc, store_code, machine_name_raw))

        bc = booth_code(mc, booth_num)
        new_booths.append((bc, mc, store_code, booth_num))

    return new_machines, new_booths


def generate_stores_sql():
    lines = ['-- ===== NEW STORES =====']
    for (code, name, stype) in NEW_STORES:
        lines.append(
            f"INSERT INTO stores (store_code, store_name, store_name_official, store_type, "
            f"is_active, created_at, updated_at, updated_by) VALUES "
            f"({sq(code)}, {sq(name)}, {sq(name)}, {sq(stype)}, "
            f"TRUE, NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT (store_code) DO NOTHING;"
        )
    return '\n'.join(lines)


def generate_machines_sql(new_machines):
    if not new_machines:
        return '-- No new machines'
    lines = ['-- ===== NEW MACHINES =====']
    for (mc, store_code, machine_name) in new_machines:
        lines.append(
            f"INSERT INTO machines (machine_code, store_code, machine_name, play_price, "
            f"is_active, operator_id, created_at, updated_at, updated_by) VALUES "
            f"({sq(mc)}, {sq(store_code)}, {sq(machine_name)}, 100, "
            f"TRUE, {sq(DEFAULT_OPERATOR_ID)}, NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT (machine_code) DO NOTHING;"
        )
    return '\n'.join(lines)


def generate_booths_sql(new_booths):
    if not new_booths:
        return '-- No new booths'
    lines = ['-- ===== NEW BOOTHS =====']
    seen_bc = set()
    for (bc, mc, store_code, booth_num) in new_booths:
        if bc in seen_bc:
            continue
        seen_bc.add(bc)
        lines.append(
            f"INSERT INTO booths (booth_code, machine_code, store_code, booth_number, "
            f"is_active, play_price, created_at, updated_at, updated_by) VALUES "
            f"({sq(bc)}, {sq(mc)}, {sq(store_code)}, {booth_num}, "
            f"TRUE, 100, NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT (booth_code) DO NOTHING;"
        )
    return '\n'.join(lines)


if __name__ == '__main__':
    donki_tuples = parse_donki_machines()
    fks_tuples = parse_fukushige_machines()
    all_tuples = donki_tuples + fks_tuples
    new_machines, new_booths = build_machine_registry(all_tuples)

    print(generate_stores_sql())
    print()
    print(generate_machines_sql(new_machines))
    print()
    print(generate_booths_sql(new_booths))
    print(f'\n-- Summary: {len(NEW_STORES)} stores, {len(new_machines)} machines, {len(new_booths)} booths')
