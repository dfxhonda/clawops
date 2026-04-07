# scripts/import_excel/phase3_zaiko.py
"""
Parses 棚卸リスト and generates prize_stocks INSERT SQL.
Requires prize_masters JSON on stdin:
  [{"prize_id": "...", "prize_name": "..."}]

Usage:
  python phase3_zaiko.py < prizes.json
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import json
import openpyxl
from config import ZAIKO_FILE, STORE_MAP, IMPORT_TAG
from helpers import fuzzy_match, new_id, sq

ZAIKO_SHEETS = ['本田', '吉井', '中島', '久留米 ', '久留米(ドンキ景品)', '飯塚', '堤', '鹿児島', '塩川', '坂本']

STORE_LOCATION_MAP = {
    'KGU01': 'KGS02', 'STS01': 'KGS03', 'MYJ01': 'KGS04',
    'KRH01': 'KGS05', 'KNY01': 'KGS06', 'KGC01': 'KGS01',
    'IIZ01': 'IZK01', 'KOS01': 'KGS01',
}
FALLBACK_LOCATION = 'KGS01'


def parse_zaiko_sheet(ws, prize_name_list, machine_registry):
    """
    Returns (stocks_list, new_prize_names_set).
    stocks_list: [{'prize_name', 'owner_type', 'owner_id', 'quantity'}]
    """
    rows = list(ws.iter_rows(values_only=True))
    stocks = []
    new_prize_names = set()

    for row in rows[3:]:
        store_name = str(row[0]).strip() if row[0] else None
        prize_name_raw = str(row[1]).strip() if row[1] else None
        machine_name_raw = str(row[2]).strip() if row[2] else None
        qty_s1 = row[4] if len(row) > 4 else None
        qty_s2 = row[5] if len(row) > 5 else None
        qty_m  = row[6] if len(row) > 6 else None

        if not prize_name_raw or prize_name_raw in ('景品名', '　', ''):
            continue
        if prize_name_raw.startswith('　') or prize_name_raw == '合計':
            continue

        store_code = STORE_MAP.get(store_name) if store_name else None

        # Fuzzy match prize
        matched = fuzzy_match(prize_name_raw, prize_name_list, cutoff=0.65)
        if not matched:
            new_prize_names.add(prize_name_raw)

        def add_stock(qty, owner_type, owner_id):
            if qty and isinstance(qty, (int, float)) and qty > 0:
                stocks.append({
                    'prize_name': prize_name_raw,
                    'matched_name': matched,
                    'owner_type': owner_type,
                    'owner_id': owner_id,
                    'quantity': int(qty),
                })

        # 機械内
        if machine_name_raw and store_code:
            from helpers import normalize
            norm_m = normalize(machine_name_raw)
            mc = machine_registry.get((store_code, norm_m))
            if mc:
                add_stock(qty_m, 'machine', mc)

        # 倉庫
        if store_code:
            loc = STORE_LOCATION_MAP.get(store_code, FALLBACK_LOCATION)
            add_stock(qty_s1, 'location', loc)
            add_stock(qty_s2, 'location', loc)

    return stocks, new_prize_names


def generate_new_prizes_sql(new_prize_names):
    if not new_prize_names:
        return '-- No new prizes'
    lines = ['-- ===== NEW PRIZE_MASTERS (unmatched from inventory) =====']
    for name in sorted(new_prize_names):
        pid = 'IMP-' + new_id()[:8]
        lines.append(
            f"INSERT INTO prize_masters (prize_id, prize_name, status, "
            f"created_at, updated_at, updated_by) VALUES "
            f"({sq(pid)}, {sq(name)}, 'unknown', NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT DO NOTHING;"
        )
    return '\n'.join(lines)


def generate_zaiko_sql(prize_id_map, stocks):
    if not stocks:
        return '-- No prize stocks'
    lines = ['-- ===== PRIZE STOCKS =====']
    for s in stocks:
        prize_name = s['matched_name'] or s['prize_name']
        pid = prize_id_map.get(prize_name)
        if not pid:
            continue
        sid = new_id()
        lines.append(
            f"INSERT INTO prize_stocks "
            f"(stock_id, prize_id, owner_type, owner_id, quantity, "
            f"last_counted_at, last_counted_by, created_at, updated_at, updated_by) VALUES "
            f"({sq(sid)}, {sq(pid)}, {sq(s['owner_type'])}, {sq(s['owner_id'])}, "
            f"{s['quantity']}, NOW(), {sq(IMPORT_TAG)}, NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT DO NOTHING;"
        )
    return '\n'.join(lines)


if __name__ == '__main__':
    prizes_raw = json.load(sys.stdin)
    prize_name_list = [p['prize_name'] for p in prizes_raw]
    prize_id_map = {p['prize_name']: p['prize_id'] for p in prizes_raw}

    from phase2_readings import build_full_machine_registry
    machine_registry = build_full_machine_registry()

    wb = openpyxl.load_workbook(ZAIKO_FILE, data_only=True)
    all_stocks = []
    all_new_prize_names = set()
    for sheet_name in ZAIKO_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        stocks, new_names = parse_zaiko_sheet(ws, prize_name_list, machine_registry)
        all_stocks.extend(stocks)
        all_new_prize_names.update(new_names)

    # Assign IDs to new prizes
    for name in all_new_prize_names:
        if name not in prize_id_map:
            prize_id_map[name] = 'IMP-' + new_id()[:8]

    print(generate_new_prizes_sql(all_new_prize_names))
    print()
    print(generate_zaiko_sql(prize_id_map, all_stocks))
    print(f'\n-- Summary: {len(all_stocks)} stock entries, {len(all_new_prize_names)} new prizes, '
          f'{sum(1 for s in all_stocks if s["matched_name"])} matched')
