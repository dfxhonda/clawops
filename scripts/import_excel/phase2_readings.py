# scripts/import_excel/phase2_readings.py
"""
Parses ドンキ日売速報 and 福重売上表, generates meter_readings INSERT SQL.

Reading creation rule:
  - Create a reading when in_meter value changes from previous date (or first date)
  - full_booth_code = '{machine_code}-B{N:02d}'
  - source = 'excel_import'
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime, date as date_class
from config import DONKI_FILE, FUKUSHIGE_FILE, STORE_MAP, DONKI_SHEETS, IMPORT_TAG
from helpers import normalize, new_id, sq, sql_ts, to_date_or_none
from phase1_masters import (EXISTING_MACHINES, EXISTING_MACHINE_MAX,
                             parse_donki_machines, parse_fukushige_machines,
                             build_machine_registry)


def extract_readings_from_block(in_vals, out_vals, dates):
    """
    in_vals/out_vals: list indexed by column number.
    dates: list of (col_idx, date).
    Returns list of {'date': date, 'in_meter': int, 'out_meter': int}.
    Only includes first date and dates where in_meter changed.
    """
    readings = []
    prev_in = None
    for col_idx, dt in dates:
        in_v = in_vals[col_idx] if col_idx < len(in_vals) else None
        out_v = out_vals[col_idx] if col_idx < len(out_vals) else None
        if not isinstance(in_v, (int, float)):
            continue
        in_v = int(in_v)
        out_v = int(out_v) if isinstance(out_v, (int, float)) else 0
        if prev_in is None or in_v != prev_in:
            readings.append({'date': dt, 'in_meter': in_v, 'out_meter': out_v})
            prev_in = in_v
    return readings


def build_full_machine_registry():
    """Build (store_code, norm_name) -> machine_code registry (existing + new)."""
    from helpers import next_machine_code
    tuples = parse_donki_machines() + parse_fukushige_machines()
    registry = dict(EXISTING_MACHINES)
    m_state = dict(EXISTING_MACHINE_MAX)
    decided = {}
    for store_code, machine_name_raw, booth_num in tuples:
        norm = normalize(machine_name_raw)
        key = (store_code, norm)
        if key in registry or key in decided:
            continue
        mc = next_machine_code(store_code, m_state)
        decided[key] = mc
        registry[key] = mc
    return registry


def parse_donki_readings(machine_registry):
    """Returns list of reading records from ドンキ日売速報."""
    wb = openpyxl.load_workbook(DONKI_FILE, data_only=True)
    all_readings = []

    for sheet_name in DONKI_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Extract date columns from row 0
        dates = []
        for col_i, v in enumerate(rows[0]):
            d = to_date_or_none(v)
            if d and col_i >= 2:
                dates.append((col_i, d))

        i = 1
        while i < len(rows):
            row = rows[i]
            if len(row) >= 2 and row[1] == '100円' and row[0] and str(row[0]).strip():
                store_name = str(row[0]).strip()
                store_code = STORE_MAP.get(store_name)
                machine_name_raw = (str(rows[i+1][0]).strip()
                                    if i+1 < len(rows) and rows[i+1][0] else None)
                booth_num = 1
                if i+2 < len(rows) and rows[i+2][0] is not None:
                    try:
                        booth_num = int(rows[i+2][0])
                    except (ValueError, TypeError):
                        booth_num = 1

                if store_code and machine_name_raw:
                    norm = normalize(machine_name_raw)
                    mc = machine_registry.get((store_code, norm))
                    if mc:
                        bc = f'{mc}-B{booth_num:02d}'
                        in_vals = list(row)
                        out_vals = list(rows[i+1]) if i+1 < len(rows) else []
                        for r in extract_readings_from_block(in_vals, out_vals, dates):
                            all_readings.append({
                                'full_booth_code': bc,
                                'read_time': r['date'],
                                'in_meter': r['in_meter'],
                                'out_meter': r['out_meter'],
                            })
                i += 12
            else:
                i += 1

    return all_readings


def parse_fukushige_readings(machine_registry):
    """Returns list of reading records from 福重売上表."""
    wb = openpyxl.load_workbook(FUKUSHIGE_FILE, data_only=True)
    ws = wb['福重']
    rows = list(ws.iter_rows(values_only=True))

    # Row 0: machine name positions (cells containing '福重')
    machine_cols = []
    seen_names = set()
    for col_i, v in enumerate(rows[0]):
        if v and '福重' in str(v):
            full_name = str(v).strip()
            if full_name in seen_names:
                continue
            seen_names.add(full_name)
            machine_name = full_name.replace('福重', '', 1).strip()
            if machine_name:
                machine_cols.append((col_i, machine_name))

    all_readings = []
    for start_col, machine_name_raw in machine_cols:
        norm = normalize(machine_name_raw)
        mc = machine_registry.get(('FKS01', norm))
        if not mc:
            continue
        bc = f'{mc}-B01'

        per_row = []
        for row in rows[2:]:  # data starts row 3 (index 2)
            if not row[0]:
                continue
            dt = to_date_or_none(row[0])
            if not dt:
                continue
            in_col = start_col
            out_col = start_col + 2
            in_v = row[in_col] if in_col < len(row) else None
            out_v = row[out_col] if out_col < len(row) else None
            if isinstance(in_v, (int, float)) and in_v > 0:
                per_row.append((dt, int(in_v), int(out_v) if isinstance(out_v, (int, float)) else 0))

        prev_in = None
        for dt, in_v, out_v in per_row:
            if prev_in is None or in_v != prev_in:
                all_readings.append({
                    'full_booth_code': bc,
                    'read_time': dt,
                    'in_meter': in_v,
                    'out_meter': out_v,
                })
                prev_in = in_v

    return all_readings


def generate_readings_sql(readings):
    if not readings:
        return '-- No readings'
    lines = ['-- ===== METER READINGS =====']
    for r in readings:
        rid = new_id()
        lines.append(
            f"INSERT INTO meter_readings "
            f"(reading_id, full_booth_code, booth_id, read_time, in_meter, out_meter, "
            f"source, created_at, created_by) VALUES "
            f"({sq(rid)}, {sq(r['full_booth_code'])}, {sq(r['full_booth_code'])}, "
            f"{sql_ts(r['read_time'])}, {r['in_meter']}, {r['out_meter']}, "
            f"{sq('excel_import')}, NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT DO NOTHING;"
        )
    return '\n'.join(lines)


if __name__ == '__main__':
    registry = build_full_machine_registry()
    donki_r = parse_donki_readings(registry)
    fks_r = parse_fukushige_readings(registry)
    all_r = donki_r + fks_r
    print(generate_readings_sql(all_r))
    print(f'\n-- Summary: {len(donki_r)} donki readings, {len(fks_r)} fukushige readings, {len(all_r)} total')
