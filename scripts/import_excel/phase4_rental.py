# scripts/import_excel/phase4_rental.py
"""
Parses レンタル売上表 3期.xlsx → billing_contracts + billing_events SQL.
Usage: python phase4_rental.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import json
import openpyxl
import calendar
from collections import defaultdict
from config import RENTAL_FILE, STORE_MAP, DEFAULT_OPERATOR_ID, IMPORT_TAG
from helpers import new_id, sq, sql_date, to_date_or_none


def parse_rental_contracts():
    """Returns list of contract dicts from 番号 sheet."""
    wb = openpyxl.load_workbook(RENTAL_FILE, data_only=True)
    ws = wb['番号']
    contracts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).startswith('R'):
            continue
        rental_no = str(row[0]).strip()
        machine_name = str(row[1]).strip() if row[1] else None
        revenue_share = float(row[2]) if row[2] and isinstance(row[2], (int, float)) else None
        store_name = str(row[4]).strip() if row[4] else None
        store_code = STORE_MAP.get(store_name)
        if not machine_name:
            continue
        if not store_code:
            # Use store_name as placeholder note
            store_code = None
        contracts.append({
            'contract_id': f'RENT-{rental_no}',
            'store_code': store_code,
            'operator_id': DEFAULT_OPERATOR_ID,
            'contract_type': 'rental',
            'revenue_share': revenue_share,
            'machine_name': machine_name,
            'rental_no': rental_no,
            'notes': f'{rental_no} {machine_name}',
            'is_active': True,
        })
    return contracts


def parse_rental_events():
    """
    Returns dict: {(store_code, month_date): {'total_sales': int, 'machines': list}}
    """
    wb = openpyxl.load_workbook(RENTAL_FILE, data_only=True)
    ws = wb['change']
    rows = list(ws.iter_rows(values_only=True))

    # Row 0: date headers at cols 2, 7, 12, ... (every 5 cols)
    month_cols = []
    for col_i, v in enumerate(rows[0]):
        d = to_date_or_none(v)
        if d and col_i >= 2:
            month_cols.append((col_i, d))

    events = defaultdict(lambda: {'total_sales': 0, 'machines': []})

    for row in rows[2:]:
        machine_name = str(row[0]).strip() if row[0] else None
        if not machine_name or machine_name in ('機械名', ''):
            continue

        for (base_col, month_date) in month_cols:
            # 5-col block: base_col=歩率, +1=設置店, +2=売上, +3=取得, +4=blank
            store_col = base_col + 1
            sale_col  = base_col + 2
            take_col  = base_col + 3

            store_name = str(row[store_col]).strip() if store_col < len(row) and row[store_col] else None
            sale_v = row[sale_col] if sale_col < len(row) else None
            take_v = row[take_col] if take_col < len(row) else None

            if not isinstance(sale_v, (int, float)) or sale_v <= 0:
                continue

            store_code = STORE_MAP.get(store_name) if store_name else None
            if not store_code:
                continue

            key = (store_code, month_date)
            events[key]['total_sales'] += int(sale_v)
            events[key]['machines'].append({
                'machine_name': machine_name,
                'sales': int(sale_v),
                'take': int(take_v) if isinstance(take_v, (int, float)) else 0,
            })

    return dict(events)


def generate_contracts_sql(contracts):
    valid = [c for c in contracts if c['store_code']]
    skipped = len(contracts) - len(valid)
    lines = [f'-- ===== BILLING CONTRACTS ({len(valid)} contracts, {skipped} skipped - unknown store) =====']
    for c in valid:
        rs = str(c['revenue_share']) if c['revenue_share'] is not None else 'NULL'
        lines.append(
            f"INSERT INTO billing_contracts "
            f"(contract_id, store_code, operator_id, contract_type, revenue_share, "
            f"is_active, notes, created_at, updated_at, updated_by) VALUES "
            f"({sq(c['contract_id'])}, {sq(c['store_code'])}, {sq(c['operator_id'])}, "
            f"{sq(c['contract_type'])}, {rs}, TRUE, "
            f"{sq(c['notes'])}, NOW(), NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT (contract_id) DO NOTHING;"
        )
    return '\n'.join(lines)


def generate_events_sql(events):
    if not events:
        return '-- No billing events'
    lines = ['-- ===== BILLING EVENTS =====']
    for (store_code, month_date), data in sorted(events.items()):
        bid = new_id()
        last_day = calendar.monthrange(month_date.year, month_date.month)[1]
        period_from = month_date.replace(day=1)
        period_to = month_date.replace(day=last_day)
        machine_json = json.dumps(data['machines'], ensure_ascii=False)
        lines.append(
            f"INSERT INTO billing_events "
            f"(billing_id, store_code, billing_date, period_from, period_to, "
            f"total_sales, status, machine_details, created_by, created_at, updated_by) VALUES "
            f"({sq(bid)}, {sq(store_code)}, {sql_date(period_to)}, "
            f"{sql_date(period_from)}, {sql_date(period_to)}, "
            f"{data['total_sales']}, 'draft', "
            f"{sq(machine_json)}, "
            f"{sq(IMPORT_TAG)}, NOW(), {sq(IMPORT_TAG)}) "
            f"ON CONFLICT DO NOTHING;"
        )
    return '\n'.join(lines)


if __name__ == '__main__':
    contracts = parse_rental_contracts()
    events = parse_rental_events()
    print(generate_contracts_sql(contracts))
    print()
    print(generate_events_sql(events))
    print(f'\n-- Summary: {len(contracts)} contracts total, {sum(1 for c in contracts if c["store_code"])} with store_code, '
          f'{len(events)} monthly billing events')
