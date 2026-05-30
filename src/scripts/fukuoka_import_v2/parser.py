#!/usr/bin/env python3
"""
Fukuoka import v2 — clawops Round Zero
=======================================

ヒアリング結果を反映した取り込み専用 parser。

仕様要点(2026-05-30 ヒロ確定):
  - R1 = 機械名(列単位)、R2 = 各機械のヘッダ(日付/100円/プライズ/投入残/補充/景品/単価/設定 …)
  - 機械の列数は機械によって変動。R2 の「日付」列でブロック分割
  - ブロック内 R2 のラベル文字列で取り込みカラムを動的検出
  - 計算カラム(差/払出/出率/払出円/合計)は捨てる
  - 集計行(最左セルが "月末合計" 等のラベル)は skip
  - 日付セルが空 → 直前の日付を継承
  - 機械名に R番号(R\\d{3,4}) があれば DFX 所有マーク
  - シートが複数ある場合は全シート集約 (1ファイル = 1店舗)

ブース構造は「列方向展開」と判断:
  - R1 で機械名が出現するブロックを基準に、R1 が空のブロックは
    「直前機械の追加ブース」として booth_number をインクリメント
  - これはヒアリングメモ D の「行方向」と食い違うため Phase 4 で要確認

Output: JSON-Lines 中間ファイル(本番投入は別スクリプトで実施)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Iterator

import openpyxl  # type: ignore


# ----------------------------------------------------------------------------
# 取り込み対象ヘッダ → 出力カラム名 のマッピング
# ----------------------------------------------------------------------------
HEADER_ALIAS: dict[str, str] = {
    "100円": "in_meter",
    "100": "in_meter",
    "プライズ": "out_meter",
    "投入残": "theoretical_stock",
    "残数": "theoretical_stock",
    "補充": "prize_restock_count",
    "景品": "prize_name",
    "景品名": "prize_name",
    "単価": "prize_cost",
    "設定": "note",
}

# 計算で再現できるので捨てるヘッダ
DROP_HEADERS: set[str] = {
    "差", "払出", "払出円", "出率", "合計", "原価率", "払出額",
}

# 集計行ラベル(日付セルにこの文字列があったら行ごと skip)
AGGREGATE_LABELS: set[str] = {
    "月末合計", "月計", "月締", "月末", "月末在庫金額", "月末在庫",
    "利益", "原価", "トータル原価", "原価合計",
    "払出合計", "単価", "景品", "トータル",
    "合計", "計", "小計", "集計", "差引",
    "棚卸", "期末", "メモ", "注意", "備考",
    "在庫金額",
}

# R番号抽出パターン (R1014, R3003, R30, R番号付きの機械名)
R_NUMBER_PATTERN = re.compile(r"R\s*(\d{2,5})")


# ----------------------------------------------------------------------------
# データクラス
# ----------------------------------------------------------------------------
@dataclass
class BlockSpec:
    """1機械ブロック(列方向の範囲)の定義"""
    date_col: int
    col_end: int  # inclusive
    header_cols: dict[str, int] = field(default_factory=dict)
    raw_machine_label: str | None = None
    machine_key: str | None = None  # 継承後の機械名
    booth_within_machine: int = 1

    @property
    def col_start(self) -> int:
        return self.date_col


@dataclass
class ParsedRow:
    file_name: str
    sheet_name: str
    store_code: str
    machine_key: str
    booth_within_machine: int
    r_number: str | None  # "R1014" 等(DFX所有マーカー)
    patrol_date: str  # ISO YYYY-MM-DD
    in_meter: float | None
    out_meter: float | None
    theoretical_stock: int | None
    prize_restock_count: int | None
    prize_name: str | None
    prize_cost: float | None
    note: str | None
    source_row: int  # Excel上の行番号(1-based)


# ----------------------------------------------------------------------------
# ヘルパー
# ----------------------------------------------------------------------------
def cell_text(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def as_date(v: Any) -> date | None:
    """セル値を date に変換。文字列の場合は集計ラベルじゃないかを呼び出し側でチェック済の前提。"""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # よくある日本語日付フォーマットを試す
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d", "%Y年%m月%d日"):
            try:
                d = datetime.strptime(s, fmt)
                # %m/%d だけだと年が 1900 になるので、現在年を使う
                if d.year == 1900:
                    d = d.replace(year=datetime.now().year)
                return d.date()
            except ValueError:
                continue
    return None


def is_aggregate_label(v: Any) -> bool:
    """日付セルの値が集計ラベルなら True"""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s:
        return False
    # 完全一致 or 部分一致(「月末在庫金額」等のロングラベルにも対応)
    if s in AGGREGATE_LABELS:
        return True
    for lbl in AGGREGATE_LABELS:
        if lbl in s:
            return True
    return False


def to_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_int(v: Any) -> int | None:
    n = to_number(v)
    if n is None:
        return None
    return int(round(n))


def extract_r_number(label: str | None) -> str | None:
    if not label:
        return None
    m = R_NUMBER_PATTERN.search(label)
    if not m:
        return None
    return "R" + m.group(1)


# ----------------------------------------------------------------------------
# ブロック検出
# ----------------------------------------------------------------------------
def detect_blocks(
    r1_row: tuple[Any, ...], r2_row: tuple[Any, ...]
) -> list[BlockSpec]:
    """R2 で「日付」がある列を起点にブロック分割。"""
    blocks: list[BlockSpec] = []
    n = max(len(r1_row), len(r2_row))
    for c in range(n):
        v = r2_row[c] if c < len(r2_row) else None
        if cell_text(v) == "日付":
            blocks.append(BlockSpec(date_col=c, col_end=n - 1))

    # col_end の確定(次ブロックの date_col - 1)
    for i, b in enumerate(blocks):
        if i + 1 < len(blocks):
            b.col_end = blocks[i + 1].date_col - 1

    # ヘッダマッピング
    for b in blocks:
        for c in range(b.col_start, b.col_end + 1):
            v = r2_row[c] if c < len(r2_row) else None
            h = cell_text(v)
            if not h:
                continue
            if h in HEADER_ALIAS:
                # 同じヘッダが2回ある場合は最初に出てきた列を採用
                b.header_cols.setdefault(HEADER_ALIAS[h], c)

    # 機械名(R1)
    prev_machine = None
    booth_counter = 0
    for b in blocks:
        pieces = []
        for c in range(b.col_start, b.col_end + 1):
            v = r1_row[c] if c < len(r1_row) else None
            t = cell_text(v)
            if t:
                pieces.append(t)
        if pieces:
            label = " | ".join(pieces)
            b.raw_machine_label = label
            prev_machine = label
            booth_counter = 1
        else:
            booth_counter += 1
        b.machine_key = prev_machine
        b.booth_within_machine = booth_counter if prev_machine else 1

    return blocks


# ----------------------------------------------------------------------------
# シート解析本体
# ----------------------------------------------------------------------------
def parse_sheet(
    ws,
    file_name: str,
    store_code: str,
) -> tuple[list[ParsedRow], dict]:
    """1シートを解析して取り込み行を返す。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], {"sheet": ws.title, "skipped": True, "reason": "rows<3"}

    r1 = rows[0]
    r2 = rows[1]
    blocks = detect_blocks(r1, r2)

    # ブロック単位で current_date を持つ
    current_date_by_block: list[date | None] = [None] * len(blocks)

    parsed: list[ParsedRow] = []
    aggregate_skip = 0
    empty_row_skip = 0
    no_date_skip = 0

    for ri in range(2, len(rows)):
        row = rows[ri]
        excel_row = ri + 1  # 1-based for human readability

        for bi, b in enumerate(blocks):
            # 日付セル
            date_cell = row[b.date_col] if b.date_col < len(row) else None
            if is_aggregate_label(date_cell):
                aggregate_skip += 1
                # 日付ラベル系の集計行は次行に持ち越さないので current_date リセットしない
                continue
            d = as_date(date_cell)
            if d is not None:
                current_date_by_block[bi] = d

            # ヘッダから値抽出
            def get(col_name: str) -> Any:
                c = b.header_cols.get(col_name)
                if c is None or c >= len(row):
                    return None
                return row[c]

            in_meter = to_number(get("in_meter"))
            out_meter = to_number(get("out_meter"))
            theoretical_stock = to_int(get("theoretical_stock"))
            prize_restock_count = to_int(get("prize_restock_count"))
            prize_name = cell_text(get("prize_name"))
            prize_cost = to_number(get("prize_cost"))
            note = cell_text(get("note"))

            # 空行判定: 主要メトリクス全て None
            non_empty = any(
                v is not None
                for v in (
                    in_meter,
                    out_meter,
                    theoretical_stock,
                    prize_restock_count,
                    prize_name,
                    prize_cost,
                    note,
                )
            )
            if not non_empty:
                empty_row_skip += 1
                continue

            # 日付がまだ確定していない場合は skip(継承する元すらない)
            if current_date_by_block[bi] is None:
                no_date_skip += 1
                continue

            r_number = extract_r_number(b.machine_key)
            parsed.append(
                ParsedRow(
                    file_name=file_name,
                    sheet_name=ws.title,
                    store_code=store_code,
                    machine_key=b.machine_key or "(unknown)",
                    booth_within_machine=b.booth_within_machine,
                    r_number=r_number,
                    patrol_date=current_date_by_block[bi].isoformat(),
                    in_meter=in_meter,
                    out_meter=out_meter,
                    theoretical_stock=theoretical_stock,
                    prize_restock_count=prize_restock_count,
                    prize_name=prize_name,
                    prize_cost=prize_cost,
                    note=note,
                    source_row=excel_row,
                )
            )

    machines_summary = []
    for b in blocks:
        machines_summary.append({
            "machine_key": b.machine_key,
            "raw_label": b.raw_machine_label,
            "booth": b.booth_within_machine,
            "r_number": extract_r_number(b.machine_key),
            "date_col": b.date_col,
            "col_end": b.col_end,
            "headers_resolved": list(b.header_cols.keys()),
        })

    stats = {
        "sheet": ws.title,
        "blocks": len(blocks),
        "rows_total": len(rows),
        "parsed_rows": len(parsed),
        "aggregate_skip": aggregate_skip,
        "empty_row_skip": empty_row_skip,
        "no_date_skip": no_date_skip,
        "machines": machines_summary,
    }
    return parsed, stats


# ----------------------------------------------------------------------------
# ファイル全体
# ----------------------------------------------------------------------------
def parse_file(path: Path, store_code: str) -> tuple[list[ParsedRow], dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_rows: list[ParsedRow] = []
    sheet_stats = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows, stats = parse_sheet(ws, path.name, store_code)
            all_rows.extend(rows)
            sheet_stats.append(stats)
    finally:
        wb.close()
    file_stat = {
        "file": path.name,
        "store_code": store_code,
        "sheets": sheet_stats,
        "total_parsed_rows": len(all_rows),
    }
    return all_rows, file_stat


# ----------------------------------------------------------------------------
# エントリポイント
# ----------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheets-dir", default="docs/sheets")
    p.add_argument("--file-map", default="src/scripts/fukuoka_import_v2/file_map.json")
    p.add_argument("--out-dir", default="src/scripts/fukuoka_import_v2/output")
    p.add_argument(
        "--mode",
        choices=["dry-run", "emit"],
        default="dry-run",
        help="dry-run=stats only / emit=書き出し含む",
    )
    p.add_argument(
        "--only", default=None, help="特定ファイル名のみ実行(ファイル名そのもの)"
    )
    args = p.parse_args()

    sheets_dir = Path(args.sheets_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(args.file_map, encoding="utf-8") as f:
        file_map = json.load(f)

    overall_stats = {
        "files": [],
        "totals": {"parsed_rows": 0, "files_processed": 0},
        "sample_rows": [],
    }
    all_parsed_rows: list[ParsedRow] = []

    targets = list(file_map.items())
    if args.only:
        targets = [(k, v) for k, v in targets if k == args.only]

    for fname, meta in targets:
        path = sheets_dir / fname
        if not path.exists():
            print(f"WARN: file not found: {path}", file=sys.stderr)
            continue
        rows, fstat = parse_file(path, meta["store_code"])
        overall_stats["files"].append(fstat)
        overall_stats["totals"]["parsed_rows"] += len(rows)
        overall_stats["totals"]["files_processed"] += 1
        all_parsed_rows.extend(rows)
        print(
            f"  {fname} -> {meta['store_code']}: {len(rows)} parsed rows, "
            f"sheets={len(fstat['sheets'])}, "
            f"blocks_total={sum(s['blocks'] for s in fstat['sheets'])}, "
            f"agg_skip={sum(s['aggregate_skip'] for s in fstat['sheets'])}"
        )

    overall_stats["sample_rows"] = [asdict(r) for r in all_parsed_rows[:20]]

    # サマリ JSON
    stats_path = out_dir / "dry_run_summary.json"
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(overall_stats, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nWrote summary: {stats_path}")

    if args.mode == "emit":
        # 全行を JSON-Lines に出力
        jsonl_path = out_dir / "parsed_rows.jsonl"
        with open(jsonl_path, "w", encoding="utf-8") as f:
            for r in all_parsed_rows:
                f.write(json.dumps(asdict(r), ensure_ascii=False, default=str) + "\n")
        print(f"Wrote rows : {jsonl_path}  ({len(all_parsed_rows)} rows)")


if __name__ == "__main__":
    main()
