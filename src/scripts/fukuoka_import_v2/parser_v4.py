#!/usr/bin/env python3
"""
Fukuoka import v4 — clawops Round Zero
=======================================

ヒロ確定仕様 v3 リファイン(2026-05-30 第3回):

修正4点:
  A) 機械 unique key を「機械名 + 設置エリア」に変更
     - "BUZZ① 1F" と "BUZZ① 2F" は別機械
     - エリア表記候補: 1F/2F/3F/B1/一階/二階/三階/エスカ/バーバー横/二階吹抜 等
  B) 100円(in_meter)列を数値型のみに絞る
     - 文字列(="月末合計"等のラベルだが to_number で 0 になるケースを含む)はskip
  C) 9999991系出荷時メーター残値は通常数値として保存(特殊扱いなし)
  D) BUZZ系は4ブース固定なので、b5以降は anomaly source として分離保存

ヒロ仕様(再掲):
  - R1 = 機械名、R2 = ヘッダ、R3以降 = 日次データ
  - 1機械=1/2/4ブース、ブースは行方向(同日付の連続行)展開
  - 同名機械が R1 に複数回 = 別ユニット
  - 「時津」「浜の町」等の店舗名のみ機械名は誤記skip
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore


# ----------------------------------------------------------------------------
# 取り込み対象ヘッダ → 出力カラム名
# ----------------------------------------------------------------------------
HEADER_ALIAS: dict[str, str] = {
    "100円": "in_meter",
    "100": "in_meter",
    "プライズ": "out_meter",
    "投入残": "theoretical_stock",
    "残数": "theoretical_stock",
    "補充": "prize_restock_count",
    "景品": "prize_name",
    "景品名": "prize_name",
    "単価": "prize_cost",
    "設定": "note",
}

# R番号(機械コード R\d{2,5})
R_NUMBER_PATTERN = re.compile(r"R\s*(\d{2,5})")

# 店舗名そのものの誤記検出
STORE_NAME_TOKENS: set[str] = {
    "時津", "浜の町", "浜町", "福江", "福重",
    "中州", "中洲", "佐賀", "唐津", "西新", "那珂川",
    "ダイキョー", "ダイキョー弥永", "ベイサイド",
    "ファミマ", "ファミマ呉服町", "ファミリーマート",
}

# エリア表記候補(機械名と分離する)
AREA_TOKENS_EXACT: set[str] = {
    "1F", "2F", "3F", "4F", "B1", "B1F", "B2",
    "1階", "2階", "3階", "4階",
    "一階", "二階", "三階", "四階", "地下",
}
AREA_TOKENS_FUZZY: list[str] = [
    "エスカ", "バーバー", "二階吹抜", "二階奥", "二階キッズ", "二階SR",
    "入口", "通路", "中央", "奥", "店内", "店外", "屋外",
]


# ----------------------------------------------------------------------------
# データクラス
# ----------------------------------------------------------------------------
@dataclass
class MachineBlock:
    machine_name: str        # 機械名のみ("BUZZ①" 等)
    area: str | None         # 設置エリア("1F" 等)、なければ None
    unit_index: int          # 同名同エリア機械の通し番号(1から)
    date_col: int
    col_end: int
    header_cols: dict[str, int] = field(default_factory=dict)
    r_number: str | None = None
    r1_raw_label: str = ""


@dataclass
class ParsedRow:
    file_name: str
    sheet_name: str
    store_code: str
    machine_name: str
    machine_area: str | None
    unit_index: int
    booth_number: int
    is_anomaly: bool         # booth>4 等の異常データ
    anomaly_reason: str | None
    r_number: str | None
    patrol_date: str
    in_meter: float | None
    out_meter: float | None
    theoretical_stock: int | None
    prize_restock_count: int | None
    prize_name: str | None
    prize_cost: float | None
    note: str | None
    source_row: int


# ----------------------------------------------------------------------------
# ヘルパー
# ----------------------------------------------------------------------------
def cell_text(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d", "%Y年%m月%d日"):
            try:
                d = datetime.strptime(s, fmt)
                if d.year == 1900:
                    d = d.replace(year=datetime.now().year)
                return d.date()
            except ValueError:
                continue
    return None


def is_strict_numeric(v: Any) -> bool:
    """ヒロ仕様: 100円列は『空 or 数値以外はskip』。
    厳密に int/float のみ true(boolは除外)。
    """
    if isinstance(v, bool):
        return False
    return isinstance(v, (int, float))


def to_number(v: Any) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_int(v: Any) -> int | None:
    n = to_number(v)
    if n is None:
        return None
    return int(round(n))


def extract_r_number(label: str | None) -> str | None:
    if not label:
        return None
    m = R_NUMBER_PATTERN.search(label)
    if not m:
        return None
    return "R" + m.group(1)


def is_store_name_only(label: str) -> bool:
    if not label:
        return True
    s = label.strip()
    return s in STORE_NAME_TOKENS


def is_area_token(label: str) -> bool:
    if not label:
        return False
    s = label.strip()
    if s in AREA_TOKENS_EXACT:
        return True
    for t in AREA_TOKENS_FUZZY:
        if t in s and len(s) <= 12:  # 機械名に "奥" が含まれることもあるので長さ制限
            return True
    return False


def normalize_machine_name_and_area(
    raw_pieces: list[str],
) -> tuple[str | None, str | None]:
    """R1 ブロック内の非空セル群を 機械名と設置エリアに分離。
    店舗名のみのピースは無視。エリアトークンはエリアへ、その他は機械名へ。
    """
    if not raw_pieces:
        return None, None
    # 店舗名除去
    non_store = [p for p in raw_pieces if not is_store_name_only(p)]
    if not non_store:
        return None, None

    machine_parts = []
    area_parts = []
    for p in non_store:
        if is_area_token(p):
            area_parts.append(p)
        else:
            machine_parts.append(p)

    if not machine_parts:
        # 全部エリアトークンだった → 機械名抽出失敗
        return None, None

    machine_name = " ".join(machine_parts).strip()
    area = " ".join(area_parts).strip() if area_parts else None
    return machine_name, area


# ----------------------------------------------------------------------------
# 機械ブロック検出
# ----------------------------------------------------------------------------
def detect_machine_blocks(
    r1_row: tuple[Any, ...], r2_row: tuple[Any, ...]
) -> list[MachineBlock]:
    n = max(len(r1_row), len(r2_row))

    date_cols: list[int] = []
    for c in range(n):
        v = r2_row[c] if c < len(r2_row) else None
        if cell_text(v) == "日付":
            date_cols.append(c)

    col_ends: list[int] = []
    for i, dc in enumerate(date_cols):
        col_ends.append(date_cols[i + 1] - 1 if i + 1 < len(date_cols) else n - 1)

    raw_blocks: list[MachineBlock] = []
    for dc, ce in zip(date_cols, col_ends):
        r1_pieces = []
        for c in range(dc, ce + 1):
            v = r1_row[c] if c < len(r1_row) else None
            t = cell_text(v)
            if t:
                r1_pieces.append(t)

        if not r1_pieces:
            continue  # R1 空ブロックは無視

        machine_name, area = normalize_machine_name_and_area(r1_pieces)
        if not machine_name:
            continue  # 店舗名のみ等 = 誤記skip

        # ヘッダ列マッピング
        header_cols: dict[str, int] = {}
        for c in range(dc, ce + 1):
            v = r2_row[c] if c < len(r2_row) else None
            h = cell_text(v)
            if not h:
                continue
            if h in HEADER_ALIAS:
                header_cols.setdefault(HEADER_ALIAS[h], c)

        raw_label = " | ".join(r1_pieces)
        r_number = extract_r_number(raw_label) or extract_r_number(machine_name)

        raw_blocks.append(
            MachineBlock(
                machine_name=machine_name,
                area=area,
                unit_index=0,
                date_col=dc,
                col_end=ce,
                header_cols=header_cols,
                r_number=r_number,
                r1_raw_label=raw_label,
            )
        )

    # 同名同エリア機械の unit_index 採番
    counts: dict[tuple[str, str | None], int] = defaultdict(int)
    for b in raw_blocks:
        key = (b.machine_name, b.area)
        counts[key] += 1
        b.unit_index = counts[key]

    return raw_blocks


# ----------------------------------------------------------------------------
# シート解析
# ----------------------------------------------------------------------------
# BUZZ系/類似機械は4ブース固定(ヒロ仕様)
MAX_NORMAL_BOOTHS = 4


def parse_sheet(
    ws, file_name: str, store_code: str
) -> tuple[list[ParsedRow], dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], {"sheet": ws.title, "skipped": True, "reason": "rows<3"}

    r1 = rows[0]
    r2 = rows[1]
    blocks = detect_machine_blocks(r1, r2)

    parsed: list[ParsedRow] = []
    stat_in_meter_empty = 0
    stat_in_meter_non_numeric = 0
    stat_no_date_skip = 0
    stat_normal_rows = 0
    stat_anomaly_rows = 0

    # R1空ブロック / 店舗名のみブロック の集計
    n = max(len(r1), len(r2))
    date_cols_all = [c for c in range(n) if cell_text(r2[c] if c < len(r2) else None) == "日付"]
    r1_empty_blocks = 0
    store_name_only_blocks = 0
    for c in date_cols_all:
        col_end = next((dc - 1 for dc in date_cols_all if dc > c), n - 1)
        r1_pieces = [cell_text(r1[cc] if cc < len(r1) else None) for cc in range(c, col_end + 1)]
        r1_pieces = [p for p in r1_pieces if p]
        if not r1_pieces:
            r1_empty_blocks += 1
            continue
        machine_name, area = normalize_machine_name_and_area(r1_pieces)
        if not machine_name:
            store_name_only_blocks += 1

    for b in blocks:
        current_date: date | None = None
        booth_in_group = 0

        for ri in range(2, len(rows)):
            row = rows[ri]
            excel_row = ri + 1
            in_meter_col = b.header_cols.get("in_meter")
            in_raw = row[in_meter_col] if in_meter_col is not None and in_meter_col < len(row) else None

            # 第一フィルタ: 100円(in_meter) が数値型でない行は skip
            if in_raw is None or (isinstance(in_raw, str) and in_raw.strip() == ""):
                stat_in_meter_empty += 1
                continue
            if not is_strict_numeric(in_raw):
                stat_in_meter_non_numeric += 1
                continue

            in_meter = float(in_raw)

            # 日付確定
            date_cell = row[b.date_col] if b.date_col < len(row) else None
            d = as_date(date_cell)
            if d is not None:
                current_date = d
                booth_in_group = 1
            else:
                if current_date is None:
                    stat_no_date_skip += 1
                    continue
                booth_in_group += 1

            def get(col_name: str) -> Any:
                c = b.header_cols.get(col_name)
                if c is None or c >= len(row):
                    return None
                return row[c]

            out_meter = to_number(get("out_meter"))
            theoretical_stock = to_int(get("theoretical_stock"))
            prize_restock_count = to_int(get("prize_restock_count"))
            prize_name = cell_text(get("prize_name"))
            prize_cost = to_number(get("prize_cost"))
            note = cell_text(get("note"))

            is_anomaly = booth_in_group > MAX_NORMAL_BOOTHS
            anomaly_reason = (
                f"booth>{MAX_NORMAL_BOOTHS} (machine={b.machine_name}, area={b.area})"
                if is_anomaly else None
            )
            if is_anomaly:
                stat_anomaly_rows += 1
            else:
                stat_normal_rows += 1

            parsed.append(
                ParsedRow(
                    file_name=file_name,
                    sheet_name=ws.title,
                    store_code=store_code,
                    machine_name=b.machine_name,
                    machine_area=b.area,
                    unit_index=b.unit_index,
                    booth_number=booth_in_group,
                    is_anomaly=is_anomaly,
                    anomaly_reason=anomaly_reason,
                    r_number=b.r_number,
                    patrol_date=current_date.isoformat(),
                    in_meter=in_meter,
                    out_meter=out_meter,
                    theoretical_stock=theoretical_stock,
                    prize_restock_count=prize_restock_count,
                    prize_name=prize_name,
                    prize_cost=prize_cost,
                    note=note,
                    source_row=excel_row,
                )
            )

    # 各機械のブース統計
    booths_per_machine: dict[tuple[str, str | None, int], int] = defaultdict(int)
    for r in parsed:
        key = (r.machine_name, r.machine_area, r.unit_index)
        if r.booth_number > booths_per_machine[key]:
            booths_per_machine[key] = r.booth_number

    machines_summary = []
    for b in blocks:
        key = (b.machine_name, b.area, b.unit_index)
        machines_summary.append({
            "machine_name": b.machine_name,
            "area": b.area,
            "unit_index": b.unit_index,
            "r_number": b.r_number,
            "raw_label": b.r1_raw_label,
            "date_col": b.date_col,
            "col_end": b.col_end,
            "headers_resolved": list(b.header_cols.keys()),
            "max_booths_seen": booths_per_machine.get(key, 0),
        })

    stats = {
        "sheet": ws.title,
        "blocks_adopted": len(blocks),
        "blocks_ignored_r1_empty": r1_empty_blocks,
        "blocks_ignored_store_name_only": store_name_only_blocks,
        "rows_total": len(rows),
        "parsed_normal_rows": stat_normal_rows,
        "parsed_anomaly_rows": stat_anomaly_rows,
        "in_meter_empty_skip": stat_in_meter_empty,
        "in_meter_non_numeric_skip": stat_in_meter_non_numeric,
        "no_date_skip": stat_no_date_skip,
        "machines": machines_summary,
    }
    return parsed, stats


def parse_file(path: Path, store_code: str) -> tuple[list[ParsedRow], dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_rows: list[ParsedRow] = []
    sheet_stats = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows, stats = parse_sheet(ws, path.name, store_code)
            all_rows.extend(rows)
            sheet_stats.append(stats)
    finally:
        wb.close()
    return all_rows, {
        "file": path.name,
        "store_code": store_code,
        "sheets": sheet_stats,
        "total_parsed_rows": len(all_rows),
        "total_normal": sum(1 for r in all_rows if not r.is_anomaly),
        "total_anomaly": sum(1 for r in all_rows if r.is_anomaly),
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheets-dir", default="docs/sheets")
    p.add_argument("--file-map", default="src/scripts/fukuoka_import_v2/file_map.json")
    p.add_argument("--out-dir", default="src/scripts/fukuoka_import_v2/output_v4")
    p.add_argument("--mode", choices=["dry-run", "emit"], default="dry-run")
    p.add_argument("--only", default=None)
    args = p.parse_args()

    sheets_dir = Path(args.sheets_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(args.file_map, encoding="utf-8") as f:
        file_map = json.load(f)

    overall_stats: dict = {"files": [], "totals": {"parsed_rows": 0, "normal_rows": 0, "anomaly_rows": 0}}
    all_parsed_rows: list[ParsedRow] = []

    targets = list(file_map.items())
    if args.only:
        targets = [(k, v) for k, v in targets if k == args.only]

    for fname, meta in targets:
        path = sheets_dir / fname
        if not path.exists():
            print(f"WARN: file not found: {path}", file=sys.stderr)
            continue
        rows, fstat = parse_file(path, meta["store_code"])
        overall_stats["files"].append(fstat)
        overall_stats["totals"]["parsed_rows"] += len(rows)
        overall_stats["totals"]["normal_rows"] += fstat["total_normal"]
        overall_stats["totals"]["anomaly_rows"] += fstat["total_anomaly"]
        all_parsed_rows.extend(rows)
        print(
            f"  {fname} -> {meta['store_code']}: "
            f"{len(rows)} rows (normal={fstat['total_normal']}, anomaly={fstat['total_anomaly']}), "
            f"machines={sum(len(s['machines']) for s in fstat['sheets'])}"
        )

    stats_path = out_dir / "dry_run_summary_v4.json"
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(overall_stats, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nWrote: {stats_path}")

    if args.mode == "emit":
        jsonl_path = out_dir / "parsed_rows_v4.jsonl"
        with open(jsonl_path, "w", encoding="utf-8") as f:
            for r in all_parsed_rows:
                f.write(json.dumps(asdict(r), ensure_ascii=False, default=str) + "\n")
        print(f"Wrote: {jsonl_path}  ({len(all_parsed_rows)} rows)")


if __name__ == "__main__":
    main()
