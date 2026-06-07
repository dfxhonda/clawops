#!/usr/bin/env python3
"""
Fukuoka import v3 — clawops Round Zero
=======================================

ヒロ確定仕様 v2 (2026-05-30 第2回):

ブース解釈の根本転換:
  - 旧 v2 解釈: 列方向に複数機械、各ブロック=1ブース
  - **新 v3 解釈**: 1機械 = 複数ブース(1/2/4)、ブースは行方向(同日付の連続行)
  - R1 に同じ機械名が複数回登場 = 別ユニット(同型機の n 台目)、各ユニットが4ブース持ち得る
  - R1 が空のブロックは無視(過去仕様のレガシーで、データは scattered)

機械名フィルタ:
  - 機械名が "時津" "浜の町" "福江" など店舗名そのものなら誤記 skip

データ抽出:
  - R1 で機械名が明示されたブロックだけ採用
  - 各ブロック内のデータ行を上から走査:
    - 日付セルに日付値あり → 新グループ開始, booth=1
    - 日付セル空 → 直前日付継承, booth インクリメント
    - 日付セルに集計ラベル → skip
  - そのブロックの 100円/プライズ/投入残/補充/景品/単価/設定 を抽出

ユニット番号:
  - 同シート内に同じ機械名のブロックが N 個あれば U1, U2, ... と振る
  - 1個しかなければ unit=1 で固定
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore


# ----------------------------------------------------------------------------
# 取り込み対象ヘッダ → 出力カラム名 のマッピング
# ----------------------------------------------------------------------------
HEADER_ALIAS: dict[str, str] = {
    "100円": "in_meter",
    "100": "in_meter",
    "プライズ": "out_meter",
    "投入残": "theoretical_stock",
    "残数": "theoretical_stock",
    "補充": "prize_restock_count",
    "景品": "prize_name",
    "景品名": "prize_name",
    "単価": "prize_cost",
    "設定": "note",
}

AGGREGATE_LABELS: set[str] = {
    "月末合計", "月計", "月締", "月末", "月末在庫金額", "月末在庫",
    "利益", "原価", "トータル原価", "原価合計",
    "払出合計", "単価", "景品", "トータル",
    "合計", "計", "小計", "集計", "差引",
    "棚卸", "期末", "メモ", "注意", "備考",
    "在庫金額",
}

# R番号抽出(R + 数字2〜5桁)
R_NUMBER_PATTERN = re.compile(r"R\s*(\d{2,5})")

# 店舗名そのものが機械名として記録されている誤記を弾くためのリスト
# ファイル名から推測される地名語幹を網羅
STORE_NAME_TOKENS: set[str] = {
    "時津", "浜の町", "浜町", "福江", "福重",
    "中州", "中洲", "佐賀", "唐津", "西新", "那珂川",
    "ダイキョー", "ダイキョー弥永", "ベイサイド",
    "ファミマ", "ファミマ呉服町", "ファミリーマート",
}


# ----------------------------------------------------------------------------
# データクラス
# ----------------------------------------------------------------------------
@dataclass
class MachineBlock:
    """1機械(ユニット)の列ブロック"""
    machine_name: str          # 例: "BUZZ①" / "カリーノミニSP" / "唐津R3003"
    unit_index: int            # 同名機械が複数あるときの 1..N
    date_col: int
    col_end: int
    header_cols: dict[str, int] = field(default_factory=dict)
    r_number: str | None = None
    r1_raw_label: str = ""     # R1 で機械名と並んで書かれてた付帯情報(フロア等)


@dataclass
class ParsedRow:
    file_name: str
    sheet_name: str
    store_code: str
    machine_name: str
    unit_index: int            # 同名機械内の通し番号
    booth_number: int          # 1..N
    r_number: str | None
    patrol_date: str           # ISO YYYY-MM-DD
    in_meter: float | None
    out_meter: float | None
    theoretical_stock: int | None
    prize_restock_count: int | None
    prize_name: str | None
    prize_cost: float | None
    note: str | None
    source_row: int            # Excel行番号(1-based)


# ----------------------------------------------------------------------------
# ヘルパー
# ----------------------------------------------------------------------------
def cell_text(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d", "%Y年%m月%d日"):
            try:
                d = datetime.strptime(s, fmt)
                if d.year == 1900:
                    d = d.replace(year=datetime.now().year)
                return d.date()
            except ValueError:
                continue
    return None


def is_aggregate_label(v: Any) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s:
        return False
    if s in AGGREGATE_LABELS:
        return True
    for lbl in AGGREGATE_LABELS:
        if lbl in s:
            return True
    return False


def to_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_int(v: Any) -> int | None:
    n = to_number(v)
    if n is None:
        return None
    return int(round(n))


def extract_r_number(label: str | None) -> str | None:
    if not label:
        return None
    m = R_NUMBER_PATTERN.search(label)
    if not m:
        return None
    return "R" + m.group(1)


def is_store_name_only(label: str) -> bool:
    """ラベルが店舗名そのもの(機械名要素を含まない)なら True"""
    if not label:
        return True
    s = label.strip()
    return s in STORE_NAME_TOKENS


def normalize_machine_name(raw_label: str) -> str | None:
    """R1ブロック内のセル群を結合した raw_label から、機械の代表名を抽出。
    店舗名のみのラベルはスキップ。
    """
    if not raw_label:
        return None
    # raw_label は "唐津R3003 | BUZZ①" のようにパイプ結合されている
    parts = [p.strip() for p in raw_label.split("|") if p.strip()]
    if not parts:
        return None
    # 店舗名トークンを除いた要素のうち、最後を機械名として採用
    machine_parts = [p for p in parts if not is_store_name_only(p)]
    if not machine_parts:
        # 全部店舗名だった → 機械名なし扱い (誤記の可能性)
        return None
    # 機械名側に R 番号が混じっているケース (例: "唐津R3003") もこれで吸収
    return machine_parts[-1]


# ----------------------------------------------------------------------------
# 機械ブロック検出
# ----------------------------------------------------------------------------
def detect_machine_blocks(
    r1_row: tuple[Any, ...], r2_row: tuple[Any, ...]
) -> list[MachineBlock]:
    """R2 「日付」位置でブロック切り、R1 で機械名があるブロックだけ採用。"""
    n = max(len(r1_row), len(r2_row))

    # ブロック起点(日付セル位置)
    date_cols: list[int] = []
    for c in range(n):
        v = r2_row[c] if c < len(r2_row) else None
        if cell_text(v) == "日付":
            date_cols.append(c)

    # 各ブロックの col_end
    col_ends: list[int] = []
    for i, dc in enumerate(date_cols):
        col_ends.append(date_cols[i + 1] - 1 if i + 1 < len(date_cols) else n - 1)

    # ブロックを構築(R1 が空のものは無視)
    raw_blocks = []
    for dc, ce in zip(date_cols, col_ends):
        r1_pieces = []
        for c in range(dc, ce + 1):
            v = r1_row[c] if c < len(r1_row) else None
            t = cell_text(v)
            if t:
                r1_pieces.append(t)
        raw_label = " | ".join(r1_pieces)
        if not r1_pieces:
            # R1空ブロック → 無視
            continue

        machine_name = normalize_machine_name(raw_label)
        if not machine_name:
            # 機械名抽出失敗(店舗名のみ等) → 誤記とみなして skip
            continue

        # ヘッダ列マッピング
        header_cols: dict[str, int] = {}
        for c in range(dc, ce + 1):
            v = r2_row[c] if c < len(r2_row) else None
            h = cell_text(v)
            if not h:
                continue
            if h in HEADER_ALIAS:
                header_cols.setdefault(HEADER_ALIAS[h], c)

        # R番号(機械名そのもの + raw_label全体の両方からscan)
        r_number = extract_r_number(raw_label) or extract_r_number(machine_name)

        raw_blocks.append(
            MachineBlock(
                machine_name=machine_name,
                unit_index=0,  # 後で振る
                date_col=dc,
                col_end=ce,
                header_cols=header_cols,
                r_number=r_number,
                r1_raw_label=raw_label,
            )
        )

    # 同名機械の unit_index 採番
    counts: dict[str, int] = defaultdict(int)
    for b in raw_blocks:
        counts[b.machine_name] += 1
        b.unit_index = counts[b.machine_name]

    return raw_blocks


# ----------------------------------------------------------------------------
# シート解析本体
# ----------------------------------------------------------------------------
def parse_sheet(
    ws, file_name: str, store_code: str
) -> tuple[list[ParsedRow], dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], {"sheet": ws.title, "skipped": True, "reason": "rows<3"}

    r1 = rows[0]
    r2 = rows[1]
    blocks = detect_machine_blocks(r1, r2)

    parsed: list[ParsedRow] = []
    aggregate_skip = 0
    empty_row_skip = 0
    no_date_skip = 0
    ignored_r1_empty_blocks = 0  # R1空ブロック(=無視数)を把握するために再カウント
    ignored_store_name_blocks = 0  # 店舗名のみブロック(誤記)

    # ブロック検出時のスキップ理由を統計化したいので、もう一度ループ
    n = max(len(r1), len(r2))
    date_cols_all = [c for c in range(n) if cell_text(r2[c] if c < len(r2) else None) == "日付"]
    for c in date_cols_all:
        r1v = cell_text(r1[c] if c < len(r1) else None)
        if r1v is None:
            ignored_r1_empty_blocks += 1
    # 機械名なし(店舗名のみ)ブロックの推定
    for c in date_cols_all:
        # 列範囲
        col_end = next((dc - 1 for dc in date_cols_all if dc > c), n - 1)
        r1_pieces = [cell_text(r1[cc] if cc < len(r1) else None) for cc in range(c, col_end + 1)]
        r1_pieces = [p for p in r1_pieces if p]
        if r1_pieces and normalize_machine_name(" | ".join(r1_pieces)) is None:
            ignored_store_name_blocks += 1

    # 各ブロックごとに日付グループでブース番号採番
    # 新仕様(ヒロv3 リファイン): in_meter(100円) が値を持つ行だけが「有効行」。
    #   - in_meter 空 → 集計行/空行/メモ行 全部まとめて skip
    #   - in_meter あり + 日付明示 → 新日付グループ, booth=1
    #   - in_meter あり + 日付空 → 上の日付継承, booth++
    for b in blocks:
        current_date: date | None = None
        booth_in_group = 0

        for ri in range(2, len(rows)):
            row = rows[ri]
            excel_row = ri + 1
            date_cell = row[b.date_col] if b.date_col < len(row) else None

            def get(col_name: str) -> Any:
                c = b.header_cols.get(col_name)
                if c is None or c >= len(row):
                    return None
                return row[c]

            in_meter = to_number(get("in_meter"))
            # 第一フィルタ: in_meter に有効値ない行は無条件 skip
            if in_meter is None:
                # 集計行判定は副次的にカウントだけする(統計用)
                if is_aggregate_label(date_cell):
                    aggregate_skip += 1
                else:
                    empty_row_skip += 1
                continue

            # in_meter あり → 有効行確定
            d = as_date(date_cell)
            if d is not None:
                current_date = d
                booth_in_group = 1
            else:
                if current_date is None:
                    # in_meter はあるが、まだ一度も日付を見ていない
                    # ファイル先頭の継承元なし → skip
                    no_date_skip += 1
                    continue
                booth_in_group += 1

            out_meter = to_number(get("out_meter"))
            theoretical_stock = to_int(get("theoretical_stock"))
            prize_restock_count = to_int(get("prize_restock_count"))
            prize_name = cell_text(get("prize_name"))
            prize_cost = to_number(get("prize_cost"))
            note = cell_text(get("note"))

            parsed.append(
                ParsedRow(
                    file_name=file_name,
                    sheet_name=ws.title,
                    store_code=store_code,
                    machine_name=b.machine_name,
                    unit_index=b.unit_index,
                    booth_number=booth_in_group,
                    r_number=b.r_number,
                    patrol_date=current_date.isoformat(),
                    in_meter=in_meter,
                    out_meter=out_meter,
                    theoretical_stock=theoretical_stock,
                    prize_restock_count=prize_restock_count,
                    prize_name=prize_name,
                    prize_cost=prize_cost,
                    note=note,
                    source_row=excel_row,
                )
            )

    # 各機械のブース数(max booth_number)
    booths_per_machine: dict[tuple[str, int], int] = defaultdict(int)
    for r in parsed:
        key = (r.machine_name, r.unit_index)
        if r.booth_number > booths_per_machine[key]:
            booths_per_machine[key] = r.booth_number

    machines_summary = []
    for b in blocks:
        key = (b.machine_name, b.unit_index)
        machines_summary.append({
            "machine_name": b.machine_name,
            "unit_index": b.unit_index,
            "r_number": b.r_number,
            "raw_label": b.r1_raw_label,
            "date_col": b.date_col,
            "col_end": b.col_end,
            "headers_resolved": list(b.header_cols.keys()),
            "max_booths_seen": booths_per_machine.get(key, 0),
        })

    stats = {
        "sheet": ws.title,
        "blocks_adopted": len(blocks),
        "blocks_ignored_r1_empty": ignored_r1_empty_blocks,
        "blocks_ignored_store_name_only": ignored_store_name_blocks,
        "rows_total": len(rows),
        "parsed_rows": len(parsed),
        "aggregate_skip": aggregate_skip,
        "empty_row_skip": empty_row_skip,
        "no_date_skip": no_date_skip,
        "machines": machines_summary,
    }
    return parsed, stats


def parse_file(path: Path, store_code: str) -> tuple[list[ParsedRow], dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_rows: list[ParsedRow] = []
    sheet_stats = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows, stats = parse_sheet(ws, path.name, store_code)
            all_rows.extend(rows)
            sheet_stats.append(stats)
    finally:
        wb.close()
    file_stat = {
        "file": path.name,
        "store_code": store_code,
        "sheets": sheet_stats,
        "total_parsed_rows": len(all_rows),
    }
    return all_rows, file_stat


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheets-dir", default="docs/sheets")
    p.add_argument("--file-map", default="src/scripts/fukuoka_import_v2/file_map.json")
    p.add_argument("--out-dir", default="src/scripts/fukuoka_import_v2/output_v3")
    p.add_argument("--mode", choices=["dry-run", "emit"], default="dry-run")
    p.add_argument("--only", default=None)
    args = p.parse_args()

    sheets_dir = Path(args.sheets_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(args.file_map, encoding="utf-8") as f:
        file_map = json.load(f)

    overall_stats = {
        "files": [],
        "totals": {"parsed_rows": 0, "files_processed": 0},
        "sample_rows_per_store": {},
    }
    all_parsed_rows: list[ParsedRow] = []

    targets = list(file_map.items())
    if args.only:
        targets = [(k, v) for k, v in targets if k == args.only]

    for fname, meta in targets:
        path = sheets_dir / fname
        if not path.exists():
            print(f"WARN: file not found: {path}", file=sys.stderr)
            continue
        rows, fstat = parse_file(path, meta["store_code"])
        overall_stats["files"].append(fstat)
        overall_stats["totals"]["parsed_rows"] += len(rows)
        overall_stats["totals"]["files_processed"] += 1
        all_parsed_rows.extend(rows)
        # サンプル: 店舗ごとに最初の3行
        overall_stats["sample_rows_per_store"][meta["store_code"]] = [
            asdict(r) for r in rows[:3]
        ]
        print(
            f"  {fname} -> {meta['store_code']}: "
            f"{len(rows)} rows, "
            f"machines={sum(len(s['machines']) for s in fstat['sheets'])}, "
            f"agg_skip={sum(s['aggregate_skip'] for s in fstat['sheets'])}, "
            f"r1_empty_ignored={sum(s['blocks_ignored_r1_empty'] for s in fstat['sheets'])}, "
            f"store_name_ignored={sum(s['blocks_ignored_store_name_only'] for s in fstat['sheets'])}"
        )

    stats_path = out_dir / "dry_run_summary_v3.json"
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(overall_stats, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nWrote summary: {stats_path}")

    if args.mode == "emit":
        jsonl_path = out_dir / "parsed_rows_v3.jsonl"
        with open(jsonl_path, "w", encoding="utf-8") as f:
            for r in all_parsed_rows:
                f.write(json.dumps(asdict(r), ensure_ascii=False, default=str) + "\n")
        print(f"Wrote rows : {jsonl_path}  ({len(all_parsed_rows)} rows)")


if __name__ == "__main__":
    main()
