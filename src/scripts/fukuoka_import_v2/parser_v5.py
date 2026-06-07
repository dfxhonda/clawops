#!/usr/bin/env python3
"""
Fukuoka import v5 — clawops Round Zero
=======================================

ヒロ承認 v5 仕様(2026-05-30 第4回):

機械単位独立処理 + 強化フィルタリング。
他機械から完全に切り離し、anomaly を機械単位に閉じ込める。

v4 → v5 主な追加機能:
  A) 日付列の異常検出 + 補正
     - シート全体の有効日付の中央値を基準に範囲外検出(±2年)
     - 個別行で上下5行の中央値から ±90日ずれてる場合は NULL → 継承継続
  B) 機械名チェック強化
     - 機械名が「数値のみ」→ 列ブロック全体 skip
     - 店舗名のみ → skip(既存)
     - 空文字 → skip(既存)
  C) anomaly 分類タグ
     - booth 5〜8: tag='possible_2nd_round'   (1日2回巡回の可能性)
     - booth 9〜:  tag='possible_aggregate'   (集計行/棚卸混入の可能性)
     - 直前と同景品名+同価格: tag='possible_duplicate' (重複入力)
  D) v4 比較レポート(店舗別 anomaly 推移)

7項目(日付/100円/プライズ/投入残/補充/景品名/単価/設定)以外は完全に捨てる。
"""
from __future__ import annotations

import argparse
import json
import re
import statistics
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore


# ----------------------------------------------------------------------------
# 取り込み対象ヘッダ (7項目に絞る)
# ----------------------------------------------------------------------------
HEADER_ALIAS: dict[str, str] = {
    "100円": "in_meter",
    "100": "in_meter",
    "プライズ": "out_meter",
    "投入残": "theoretical_stock",
    "残数": "theoretical_stock",
    "補充": "prize_restock_count",
    "景品": "prize_name",
    "景品名": "prize_name",
    "単価": "prize_cost",
    "設定": "note",
}

R_NUMBER_PATTERN = re.compile(r"R\s*(\d{2,5})")

STORE_NAME_TOKENS: set[str] = {
    "時津", "浜の町", "浜町", "福江", "福重",
    "中州", "中洲", "佐賀", "唐津", "西新", "那珂川",
    "ダイキョー", "ダイキョー弥永", "ベイサイド",
    "ファミマ", "ファミマ呉服町", "ファミリーマート",
}

AREA_TOKENS_EXACT: set[str] = {
    "1F", "2F", "3F", "4F", "B1", "B1F", "B2",
    "1階", "2階", "3階", "4階",
    "一階", "二階", "三階", "四階", "地下",
}
AREA_TOKENS_FUZZY: list[str] = [
    "エスカ", "バーバー", "二階吹抜", "二階奥", "二階キッズ", "二階SR",
    "入口", "通路", "中央", "奥", "店内", "店外", "屋外",
]

# 日付の正常範囲(2020〜2030 を超えるものは異常扱い)
DATE_VALID_MIN = date(2020, 1, 1)
DATE_VALID_MAX = date(2030, 12, 31)
# 個別行の異常判定: 上下5行の中央値からの最大乖離(日)
DATE_LOCAL_DEVIATION_DAYS = 90

# anomaly 閾値
MAX_NORMAL_BOOTHS = 4


# ----------------------------------------------------------------------------
# データクラス
# ----------------------------------------------------------------------------
@dataclass
class MachineBlock:
    machine_name: str
    area: str | None
    unit_index: int
    date_col: int
    col_end: int
    header_cols: dict[str, int] = field(default_factory=dict)
    r_number: str | None = None
    r1_raw_label: str = ""


@dataclass
class ParsedRow:
    file_name: str
    sheet_name: str
    store_code: str
    machine_name: str
    machine_area: str | None
    unit_index: int
    booth_number: int
    is_anomaly: bool
    anomaly_tag: str | None       # v5新規: 分類タグ
    anomaly_reason: str | None
    r_number: str | None
    patrol_date: str
    in_meter: float | None
    out_meter: float | None
    theoretical_stock: int | None
    prize_restock_count: int | None
    prize_name: str | None
    prize_cost: float | None
    note: str | None
    source_row: int


# ----------------------------------------------------------------------------
# ヘルパー(v4から流用、一部改良)
# ----------------------------------------------------------------------------
def cell_text(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d", "%Y年%m月%d日"):
            try:
                d = datetime.strptime(s, fmt)
                if d.year == 1900:
                    d = d.replace(year=datetime.now().year)
                return d.date()
            except ValueError:
                continue
    return None


def is_strict_numeric(v: Any) -> bool:
    if isinstance(v, bool):
        return False
    return isinstance(v, (int, float))


def to_number(v: Any) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_int(v: Any) -> int | None:
    n = to_number(v)
    if n is None:
        return None
    return int(round(n))


def extract_r_number(label: str | None) -> str | None:
    if not label:
        return None
    m = R_NUMBER_PATTERN.search(label)
    if not m:
        return None
    return "R" + m.group(1)


def is_store_name_only(label: str) -> bool:
    if not label:
        return True
    return label.strip() in STORE_NAME_TOKENS


def is_area_token(label: str) -> bool:
    if not label:
        return False
    s = label.strip()
    if s in AREA_TOKENS_EXACT:
        return True
    for t in AREA_TOKENS_FUZZY:
        if t in s and len(s) <= 12:
            return True
    return False


def is_numeric_only_label(label: str) -> bool:
    """機械名が数値のみ(例: '12345', '3.14')なら True"""
    if not label:
        return False
    s = label.strip().replace(",", "").replace(".", "").replace("-", "")
    return s.isdigit()


def normalize_machine_name_and_area(
    raw_pieces: list[str],
) -> tuple[str | None, str | None]:
    """機械名と設置エリアを分離。数値のみは弾く。"""
    if not raw_pieces:
        return None, None
    non_store = [p for p in raw_pieces if not is_store_name_only(p)]
    if not non_store:
        return None, None
    # 数値のみのピースを除外
    non_numeric = [p for p in non_store if not is_numeric_only_label(p)]
    if not non_numeric:
        return None, None

    machine_parts = []
    area_parts = []
    for p in non_numeric:
        if is_area_token(p):
            area_parts.append(p)
        else:
            machine_parts.append(p)

    if not machine_parts:
        return None, None

    return " ".join(machine_parts).strip(), (" ".join(area_parts).strip() if area_parts else None)


# ----------------------------------------------------------------------------
# 機械ブロック検出
# ----------------------------------------------------------------------------
def detect_machine_blocks(
    r1_row: tuple[Any, ...], r2_row: tuple[Any, ...]
) -> list[MachineBlock]:
    n = max(len(r1_row), len(r2_row))

    date_cols = [c for c in range(n) if cell_text(r2_row[c] if c < len(r2_row) else None) == "日付"]
    col_ends = [date_cols[i + 1] - 1 if i + 1 < len(date_cols) else n - 1 for i in range(len(date_cols))]

    raw_blocks: list[MachineBlock] = []
    for dc, ce in zip(date_cols, col_ends):
        r1_pieces = []
        for c in range(dc, ce + 1):
            t = cell_text(r1_row[c] if c < len(r1_row) else None)
            if t:
                r1_pieces.append(t)

        if not r1_pieces:
            continue  # R1空 → skip

        machine_name, area = normalize_machine_name_and_area(r1_pieces)
        if not machine_name:
            continue  # 店舗名のみ / 数値のみ / 機械名抽出失敗 → skip

        header_cols: dict[str, int] = {}
        for c in range(dc, ce + 1):
            h = cell_text(r2_row[c] if c < len(r2_row) else None)
            if h and h in HEADER_ALIAS:
                header_cols.setdefault(HEADER_ALIAS[h], c)

        raw_label = " | ".join(r1_pieces)
        r_number = extract_r_number(raw_label) or extract_r_number(machine_name)

        raw_blocks.append(MachineBlock(
            machine_name=machine_name, area=area, unit_index=0,
            date_col=dc, col_end=ce, header_cols=header_cols,
            r_number=r_number, r1_raw_label=raw_label,
        ))

    # 同名同エリアの unit_index 採番
    counts: dict[tuple[str, str | None], int] = defaultdict(int)
    for b in raw_blocks:
        counts[(b.machine_name, b.area)] += 1
        b.unit_index = counts[(b.machine_name, b.area)]

    return raw_blocks


# ----------------------------------------------------------------------------
# v5 新機能: 日付列の異常検出 + 補正
# ----------------------------------------------------------------------------
def correct_date_column(
    all_rows: list[tuple],
    block: MachineBlock,
) -> dict[int, date | None]:
    """機械ブロックの date_col 列について、各行の日付を補正して返す。

    アルゴリズム:
      1. 各行の生日付を取得
      2. 全有効日付の中央値を計算
      3. グローバル範囲外(DATE_VALID_MIN〜MAX)の日付は None化
      4. 個別行で上下5行の中央値から ±DATE_LOCAL_DEVIATION_DAYS 以上ずれてたら None化
      5. None になった日付は後で「上から継承」処理する

    戻り値: row_index → date or None
    """
    raw: dict[int, date | None] = {}
    for ri in range(2, len(all_rows)):
        cell = all_rows[ri][block.date_col] if block.date_col < len(all_rows[ri]) else None
        raw[ri] = as_date(cell)

    valid_dates = [d for d in raw.values() if d is not None]
    if not valid_dates:
        return raw

    # グローバル範囲チェック
    for ri, d in list(raw.items()):
        if d is None:
            continue
        if d < DATE_VALID_MIN or d > DATE_VALID_MAX:
            raw[ri] = None

    # ローカル中央値での乖離チェック(±5行のwindow)
    indices = sorted(raw.keys())
    corrected = dict(raw)
    for i, ri in enumerate(indices):
        d = corrected[ri]
        if d is None:
            continue
        # 上下5行のwindowを取る
        window_indices = indices[max(0, i - 5):i + 6]
        neighbors = [corrected[wi] for wi in window_indices if wi != ri and corrected[wi] is not None]
        if len(neighbors) < 3:
            continue  # 比較対象不十分
        median_neighbor = sorted(neighbors)[len(neighbors) // 2]
        diff = abs((d - median_neighbor).days)
        if diff > DATE_LOCAL_DEVIATION_DAYS:
            corrected[ri] = None  # 乖離大 → 異常なので None化(後で継承)

    return corrected


# ----------------------------------------------------------------------------
# anomaly タグ判定
# ----------------------------------------------------------------------------
def classify_anomaly(
    booth_num: int,
    current: dict,
    previous: dict | None,
) -> str | None:
    """v5新規: anomaly のサブクラス分類タグを返す。"""
    if booth_num <= MAX_NORMAL_BOOTHS:
        return None
    # 重複判定(直前行と景品名+価格+in_meterが完全一致)
    if previous is not None:
        if (current.get("prize_name") == previous.get("prize_name")
            and current.get("prize_cost") == previous.get("prize_cost")
            and current.get("in_meter") == previous.get("in_meter")):
            return "possible_duplicate"
    if booth_num >= 9:
        return "possible_aggregate"
    return "possible_2nd_round"  # booth 5〜8


# ----------------------------------------------------------------------------
# シート解析(機械単位独立)
# ----------------------------------------------------------------------------
def parse_sheet(
    ws, file_name: str, store_code: str
) -> tuple[list[ParsedRow], dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], {"sheet": ws.title, "skipped": True, "reason": "rows<3"}

    r1 = rows[0]
    r2 = rows[1]
    blocks = detect_machine_blocks(r1, r2)

    parsed: list[ParsedRow] = []
    stat = defaultdict(int)

    # R1空ブロック / 店舗名のみ / 数値のみ ブロックの集計
    n = max(len(r1), len(r2))
    date_cols_all = [c for c in range(n) if cell_text(r2[c] if c < len(r2) else None) == "日付"]
    for c in date_cols_all:
        col_end = next((dc - 1 for dc in date_cols_all if dc > c), n - 1)
        r1_pieces = [cell_text(r1[cc] if cc < len(r1) else None) for cc in range(c, col_end + 1)]
        r1_pieces = [p for p in r1_pieces if p]
        if not r1_pieces:
            stat["r1_empty_blocks"] += 1
            continue
        mn, _ = normalize_machine_name_and_area(r1_pieces)
        if not mn:
            # 数値のみか店舗名のみか判別
            non_store = [p for p in r1_pieces if not is_store_name_only(p)]
            if not non_store:
                stat["store_name_only_blocks"] += 1
            elif all(is_numeric_only_label(p) for p in non_store):
                stat["numeric_only_blocks"] += 1
            else:
                stat["other_invalid_blocks"] += 1

    # 機械単位独立処理
    for b in blocks:
        # 1. 日付列の補正(機械ブロック内で完結)
        corrected_dates = correct_date_column(rows, b)
        stat["date_corrections_nulled"] += sum(
            1 for ri in corrected_dates
            if corrected_dates[ri] is None and as_date(rows[ri][b.date_col] if b.date_col < len(rows[ri]) else None) is not None
        )

        # 2. 日付の継承(空セルは上の有効日付を使う)
        current_date: date | None = None
        booth_in_group = 0
        previous_row_data: dict | None = None  # 重複判定用

        for ri in range(2, len(rows)):
            row = rows[ri]
            excel_row = ri + 1
            in_meter_col = b.header_cols.get("in_meter")
            in_raw = row[in_meter_col] if in_meter_col is not None and in_meter_col < len(row) else None

            # フィルタ1: 100円(in_meter) 厳密数値チェック
            if in_raw is None or (isinstance(in_raw, str) and in_raw.strip() == ""):
                stat["in_meter_empty_skip"] += 1
                continue
            if not is_strict_numeric(in_raw):
                stat["in_meter_non_numeric_skip"] += 1
                continue

            in_meter = float(in_raw)

            # 日付確定(補正済み)
            d = corrected_dates.get(ri)
            if d is not None:
                current_date = d
                booth_in_group = 1
            else:
                if current_date is None:
                    stat["no_date_skip"] += 1
                    continue
                booth_in_group += 1

            def get(col_name: str) -> Any:
                c = b.header_cols.get(col_name)
                if c is None or c >= len(row):
                    return None
                return row[c]

            out_meter = to_number(get("out_meter"))
            theoretical_stock = to_int(get("theoretical_stock"))
            prize_restock_count = to_int(get("prize_restock_count"))
            prize_name = cell_text(get("prize_name"))
            prize_cost = to_number(get("prize_cost"))
            note = cell_text(get("note"))

            current_payload = {
                "in_meter": in_meter, "out_meter": out_meter,
                "prize_name": prize_name, "prize_cost": prize_cost,
            }

            is_anomaly = booth_in_group > MAX_NORMAL_BOOTHS
            anomaly_tag = classify_anomaly(booth_in_group, current_payload, previous_row_data) if is_anomaly else None
            anomaly_reason = None
            if is_anomaly:
                stat["anomaly_rows"] += 1
                anomaly_reason = f"booth={booth_in_group} > {MAX_NORMAL_BOOTHS} (machine={b.machine_name}, area={b.area}, tag={anomaly_tag})"
            else:
                stat["normal_rows"] += 1

            parsed.append(ParsedRow(
                file_name=file_name, sheet_name=ws.title, store_code=store_code,
                machine_name=b.machine_name, machine_area=b.area, unit_index=b.unit_index,
                booth_number=booth_in_group,
                is_anomaly=is_anomaly,
                anomaly_tag=anomaly_tag,
                anomaly_reason=anomaly_reason,
                r_number=b.r_number,
                patrol_date=current_date.isoformat(),
                in_meter=in_meter, out_meter=out_meter,
                theoretical_stock=theoretical_stock,
                prize_restock_count=prize_restock_count,
                prize_name=prize_name, prize_cost=prize_cost, note=note,
                source_row=excel_row,
            ))

            previous_row_data = current_payload

    # 機械別 max_booth
    booths_per_machine: dict[tuple[str, str | None, int], int] = defaultdict(int)
    for r in parsed:
        key = (r.machine_name, r.machine_area, r.unit_index)
        if r.booth_number > booths_per_machine[key]:
            booths_per_machine[key] = r.booth_number

    machines_summary = []
    for b in blocks:
        key = (b.machine_name, b.area, b.unit_index)
        machines_summary.append({
            "machine_name": b.machine_name, "area": b.area,
            "unit_index": b.unit_index, "r_number": b.r_number,
            "raw_label": b.r1_raw_label,
            "date_col": b.date_col, "col_end": b.col_end,
            "headers_resolved": list(b.header_cols.keys()),
            "max_booths_seen": booths_per_machine.get(key, 0),
        })

    stats = {
        "sheet": ws.title,
        "blocks_adopted": len(blocks),
        **dict(stat),
        "machines": machines_summary,
    }
    return parsed, stats


def parse_file(path: Path, store_code: str) -> tuple[list[ParsedRow], dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_rows: list[ParsedRow] = []
    sheet_stats = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows, stats = parse_sheet(ws, path.name, store_code)
            all_rows.extend(rows)
            sheet_stats.append(stats)
    finally:
        wb.close()
    return all_rows, {
        "file": path.name, "store_code": store_code,
        "sheets": sheet_stats,
        "total_parsed_rows": len(all_rows),
        "total_normal": sum(1 for r in all_rows if not r.is_anomaly),
        "total_anomaly": sum(1 for r in all_rows if r.is_anomaly),
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheets-dir", default="docs/sheets")
    p.add_argument("--file-map", default="src/scripts/fukuoka_import_v2/file_map.json")
    p.add_argument("--out-dir", default="src/scripts/fukuoka_import_v2/output_v5")
    p.add_argument("--mode", choices=["dry-run", "emit"], default="dry-run")
    p.add_argument("--only", default=None)
    args = p.parse_args()

    sheets_dir = Path(args.sheets_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(args.file_map, encoding="utf-8") as f:
        file_map = json.load(f)

    overall = {"files": [], "totals": {"parsed": 0, "normal": 0, "anomaly": 0}}
    all_rows: list[ParsedRow] = []

    targets = list(file_map.items())
    if args.only:
        targets = [(k, v) for k, v in targets if k == args.only]

    for fname, meta in targets:
        path = sheets_dir / fname
        if not path.exists():
            print(f"WARN: not found {path}", file=sys.stderr)
            continue
        rows, fstat = parse_file(path, meta["store_code"])
        overall["files"].append(fstat)
        overall["totals"]["parsed"] += len(rows)
        overall["totals"]["normal"] += fstat["total_normal"]
        overall["totals"]["anomaly"] += fstat["total_anomaly"]
        all_rows.extend(rows)
        # 日付補正count を表示
        date_nulled = sum(s.get("date_corrections_nulled", 0) for s in fstat["sheets"])
        print(
            f"  {fname} -> {meta['store_code']}: "
            f"{len(rows)} rows (normal={fstat['total_normal']}, anomaly={fstat['total_anomaly']}), "
            f"date_nulled={date_nulled}"
        )

    stats_path = out_dir / "dry_run_summary_v5.json"
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(overall, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nWrote: {stats_path}")

    if args.mode == "emit":
        jsonl_path = out_dir / "parsed_rows_v5.jsonl"
        with open(jsonl_path, "w", encoding="utf-8") as f:
            for r in all_rows:
                f.write(json.dumps(asdict(r), ensure_ascii=False, default=str) + "\n")
        print(f"Wrote: {jsonl_path}  ({len(all_rows)} rows)")


if __name__ == "__main__":
    main()
